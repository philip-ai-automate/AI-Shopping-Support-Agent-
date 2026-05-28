"""
One-shot import: PhiXtra_5000_Phone_Catalogue.xlsx → PostgreSQL
Reads phone_brands and phones_master sheets, upserts into phone_brands and phone_catalogue.
"""
import os, sys
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "api-key-manager/.env"))

XLSX = "/home/profitbuyz.com/PhiXtra_5000_Phone_Catalogue.xlsx"


def get_conn():
    return psycopg2.connect(
        host=os.getenv("PG_HOST", "localhost"),
        port=int(os.getenv("PG_PORT", "5432")),
        user=os.getenv("PG_USER"),
        password=os.getenv("PG_PASSWORD"),
        dbname=os.getenv("PG_DB"),
    )


def _val(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    return v


def _int(v):
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _float(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def import_brands(ws, cur):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    inserted = 0
    for row in rows:
        brand, priority = row[0], row[1]
        if not brand:
            continue
        cur.execute(
            """
            INSERT INTO phone_brands (brand, priority)
            VALUES (%s, %s)
            ON CONFLICT (brand) DO UPDATE SET
                priority   = EXCLUDED.priority,
                updated_at = NOW()
            """,
            (str(brand).strip(), str(priority).strip() if priority else "Medium"),
        )
        inserted += 1
    print(f"  Brands upserted: {inserted}")


def import_phones(ws, cur):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    inserted = 0
    for row in rows:
        (product_id, brand, model_name, variant_name, release_year, price_category,
         network_type, screen_size_inches, display_type, refresh_rate_hz, screen_resolution,
         chipset_model, ram, storage, battery_capacity_mah, fast_charging_watts,
         rear_camera_main_mp, front_camera_mp, video_recording, gaming_rating,
         battery_performance, camera_quality_rating, wifi_version, bluetooth_version,
         nfc, body_material, water_resistance, fingerprint_type, available_colors,
         nigeria_market_price_naira, best_for, search_intent_tags,
         ai_summary, ai_sales_pitch) = row[:34]

        if not product_id:
            continue

        cur.execute(
            """
            INSERT INTO phone_catalogue (
                product_id, brand, model_name, variant_name, release_year,
                price_category, network_type, screen_size_inches, display_type,
                refresh_rate_hz, screen_resolution, chipset_model, ram, storage,
                battery_capacity_mah, fast_charging_watts, rear_camera_main_mp,
                front_camera_mp, video_recording, gaming_rating, battery_performance,
                camera_quality_rating, wifi_version, bluetooth_version, nfc,
                body_material, water_resistance, fingerprint_type, available_colors,
                nigeria_market_price_naira, best_for, search_intent_tags,
                ai_summary, ai_sales_pitch
            ) VALUES (
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,
                %s,%s
            )
            ON CONFLICT (product_id) DO UPDATE SET
                brand                    = EXCLUDED.brand,
                model_name               = EXCLUDED.model_name,
                variant_name             = EXCLUDED.variant_name,
                release_year             = EXCLUDED.release_year,
                price_category           = EXCLUDED.price_category,
                network_type             = EXCLUDED.network_type,
                screen_size_inches       = EXCLUDED.screen_size_inches,
                display_type             = EXCLUDED.display_type,
                refresh_rate_hz          = EXCLUDED.refresh_rate_hz,
                screen_resolution        = EXCLUDED.screen_resolution,
                chipset_model            = EXCLUDED.chipset_model,
                ram                      = EXCLUDED.ram,
                storage                  = EXCLUDED.storage,
                battery_capacity_mah     = EXCLUDED.battery_capacity_mah,
                fast_charging_watts      = EXCLUDED.fast_charging_watts,
                rear_camera_main_mp      = EXCLUDED.rear_camera_main_mp,
                front_camera_mp          = EXCLUDED.front_camera_mp,
                video_recording          = EXCLUDED.video_recording,
                gaming_rating            = EXCLUDED.gaming_rating,
                battery_performance      = EXCLUDED.battery_performance,
                camera_quality_rating    = EXCLUDED.camera_quality_rating,
                wifi_version             = EXCLUDED.wifi_version,
                bluetooth_version        = EXCLUDED.bluetooth_version,
                nfc                      = EXCLUDED.nfc,
                body_material            = EXCLUDED.body_material,
                water_resistance         = EXCLUDED.water_resistance,
                fingerprint_type         = EXCLUDED.fingerprint_type,
                available_colors         = EXCLUDED.available_colors,
                nigeria_market_price_naira = EXCLUDED.nigeria_market_price_naira,
                best_for                 = EXCLUDED.best_for,
                search_intent_tags       = EXCLUDED.search_intent_tags,
                ai_summary               = EXCLUDED.ai_summary,
                ai_sales_pitch           = EXCLUDED.ai_sales_pitch,
                updated_at               = NOW()
            """,
            (
                str(product_id).strip(),
                _val(brand), _val(model_name), _val(variant_name), _int(release_year),
                _val(price_category), _val(network_type), _float(screen_size_inches), _val(display_type),
                _int(refresh_rate_hz), _val(screen_resolution), _val(chipset_model), _val(ram), _val(storage),
                _int(battery_capacity_mah), _int(fast_charging_watts), _int(rear_camera_main_mp),
                _int(front_camera_mp), _val(video_recording), _val(gaming_rating), _val(battery_performance),
                _val(camera_quality_rating), _val(wifi_version), _val(bluetooth_version), _val(nfc),
                _val(body_material), _val(water_resistance), _val(fingerprint_type), _val(available_colors),
                _float(nigeria_market_price_naira), _val(best_for), _val(search_intent_tags),
                _val(ai_summary), _val(ai_sales_pitch),
            ),
        )
        inserted += 1
        if inserted % 500 == 0:
            print(f"  ... {inserted} phones imported")

    print(f"  Phones upserted: {inserted}")


def main():
    print(f"Opening {XLSX} ...")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    conn = get_conn()
    cur = conn.cursor()

    print("Importing brands ...")
    import_brands(wb["brands"], cur)

    print("Importing phones ...")
    import_phones(wb["phones_master"], cur)

    conn.commit()
    cur.close()
    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
