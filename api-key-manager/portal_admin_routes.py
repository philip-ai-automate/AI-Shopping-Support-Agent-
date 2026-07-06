"""
portal_admin_routes.py  — Phase 1 admin portal (/admin/*)
admin_users table and db.py are UNCHANGED. app.py (keys.phixtra.com) is UNTOUCHED.
"""
import psycopg2
import psycopg2.extras
import psycopg2.errors
import os
import secrets
import string
import json as _json
import csv
import io
import bcrypt
from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, flash, send_file, Response)

from db import get_db_connection, insert_audit_log
from lead_pipeline import (STAGE_ORDER, STAGE_LABELS, STAGE_DESCRIPTIONS,
                            next_stage, record_stage_change, get_stage_history)
from portal_utils import (money_fmt, tokens_to_credits, credits_to_tokens,
                          send_email_with_attachment, TUTORIAL_VIDEOS)

portal_admin_bp = Blueprint("portal_admin", __name__)

# Multi-product ambassador program — mirrors ambassador_routes.PRODUCT_CONFIG labels.
PRODUCT_LABELS = {"portal": "Portal (Merchant)", "school": "School", "estate": "Real Estate"}
PRODUCT_ICONS  = {"portal": "🛍️", "school": "🏫", "estate": "🏠"}
LEAD_LINK_COL  = {"portal": "tenant_id", "school": "school_id", "estate": "estate_tenant_id"}
LEAD_REF_TABLE = {"portal": "tenants",   "school": "school_profiles", "estate": "re_tenants"}


def _admin_logged_in() -> bool:
    return session.get("portal_admin_logged_in") is True

def _require_admin():
    if not _admin_logged_in():
        return redirect(url_for("portal_admin.login"))
    return None

def _admin_user() -> str:
    return session.get("portal_admin_username") or "admin"


# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("portal/admin_login.html")

    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM admin_users WHERE username=%s", (username,))
    admin = cur.fetchone()
    cur.close(); conn.close()

    # Plain-text check — same as app.py (admin_users.password is plaintext by design)
    if admin and password == admin.get("password"):
        session["portal_admin_logged_in"]  = True
        session["portal_admin_username"]   = username
        return redirect(url_for("portal_admin.customers"))

    flash("Invalid admin login.", "danger")
    return redirect(url_for("portal_admin.login"))


@portal_admin_bp.route("/logout")
def logout():
    session.pop("portal_admin_logged_in", None)
    session.pop("portal_admin_username",  None)
    session.pop("impersonate_customer_id", None)
    return redirect(url_for("portal_admin.login"))


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOMER MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/customers")
def customers():
    r = _require_admin()
    if r: return r

    q = (request.args.get("q") or "").strip().lower()
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if q:
        cur.execute("""
            SELECT c.id, c.first_name, c.last_name, c.email, c.phone_number,
                   c.email_verified, c.is_active, c.created_at,
                   t.id AS tenant_id, t.name AS tenant_name, t.domain,
                   COALESCE(tb.token_balance,0) AS token_balance
            FROM customers c
            JOIN tenants t ON t.id=c.tenant_id
            LEFT JOIN tenant_balances tb ON tb.tenant_id=t.id
            WHERE t.is_demo = FALSE
              AND (LOWER(c.email) LIKE %s
               OR LOWER(t.domain) LIKE %s
               OR LOWER(t.name) LIKE %s
               OR LOWER(CONCAT(COALESCE(c.first_name,''),' ',COALESCE(c.last_name,''))) LIKE %s)
            ORDER BY c.created_at DESC LIMIT 300""",
            (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"))
    else:
        cur.execute("""
            SELECT c.id, c.first_name, c.last_name, c.email, c.phone_number,
                   c.email_verified, c.is_active, c.created_at,
                   t.id AS tenant_id, t.name AS tenant_name, t.domain,
                   COALESCE(tb.token_balance,0) AS token_balance
            FROM customers c
            JOIN tenants t ON t.id=c.tenant_id
            LEFT JOIN tenant_balances tb ON tb.tenant_id=t.id
            WHERE t.is_demo = FALSE
            ORDER BY c.created_at DESC LIMIT 300""")

    rows = cur.fetchall() or []
    cur.close(); conn.close()

    for row in rows:
        row["balance_credits"] = tokens_to_credits(int(row.get("token_balance") or 0))
        fn = (row.get("first_name") or "").strip()
        ln = (row.get("last_name")  or "").strip()
        row["full_name"] = f"{fn} {ln}".strip() or "—"

    return render_template("portal/admin_customers.html", customers=rows, q=q)


@portal_admin_bp.route("/customers/<int:customer_id>")
def customer_detail(customer_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # NOTE: t.features is NOT included in this JOIN.
    # Fetching it here caused the 500 error because the `features` column
    # may not exist in the database yet (migration may not have run).
    # We fetch it separately below with a safe try/except instead.
    cur.execute("""
        SELECT c.*, t.name AS tenant_name, t.domain,
               COALESCE(tb.token_balance,0) AS token_balance,
               t.plan_id, t.billing_cycle, t.plan_period_start,
               COALESCE(p.slug,'free') AS plan_slug,
               COALESCE(p.name,'Free') AS plan_name,
               COALESCE(p.ai_messages_limit,100) AS ai_messages_limit
        FROM customers c
        JOIN tenants t ON t.id=c.tenant_id
        LEFT JOIN tenant_balances tb ON tb.tenant_id=t.id
        LEFT JOIN plans p ON p.id = t.plan_id
        WHERE c.id=%s""", (customer_id,))
    customer = cur.fetchone()
    if not customer:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(customer["tenant_id"])

    # Fetch tenant features safely — works even if the features column
    # does not exist yet in the database (e.g. migration not run).
    # If anything goes wrong the page still loads; features just show as off.
    tenant_features = {}
    try:
        cur.execute("SELECT features FROM tenants WHERE id=%s", (tenant_id,))
        feat_row = cur.fetchone() or {}
        raw_features = feat_row.get("features")
        if raw_features:
            if isinstance(raw_features, str):
                tenant_features = _json.loads(raw_features)
            elif isinstance(raw_features, dict):
                tenant_features = raw_features
    except Exception:
        tenant_features = {}

    onboarding_other = None
    try:
        cur.execute("SELECT onboarding_other FROM tenants WHERE id=%s", (tenant_id,))
        ob_row = cur.fetchone() or {}
        onboarding_other = ob_row.get("onboarding_other") or None
    except Exception:
        onboarding_other = None

    cur.execute("""
        SELECT id, website, key_type, is_active, token_limit, tokens_used,
               trial_activated_at, trial_expires_at, created_at
        FROM api_keys WHERE tenant_id=%s ORDER BY created_at DESC""", (tenant_id,))
    keys = cur.fetchall() or []

    cur.execute("""
        SELECT id, invoice_number, credits, amount_pence, vat_pence, currency, status, created_at
        FROM invoices WHERE tenant_id=%s ORDER BY created_at DESC LIMIT 50""", (tenant_id,))
    invs = cur.fetchall() or []

    cur.execute("""
        SELECT action, website, key_type, api_key_last4, details, created_at, admin_username
        FROM audit_logs WHERE tenant_id=%s ORDER BY created_at DESC LIMIT 100""", (tenant_id,))
    audit = cur.fetchall() or []

    # Document counts per type for the index management panel
    doc_counts = {}
    try:
        cur.execute("""
            SELECT type, COUNT(*) AS cnt
            FROM documents
            WHERE tenant_id=%s
            GROUP BY type
            ORDER BY type""", (tenant_id,))
        for dc_row in cur.fetchall():
            doc_counts[dc_row["type"]] = int(dc_row["cnt"])
    except Exception:
        doc_counts = {}

    # System prompt — prefer active agent, fall back to tenant row
    tenant_system_prompt = ""
    try:
        cur.execute("""
            SELECT COALESCE(ta.system_prompt, t.system_prompt) AS system_prompt
            FROM tenants t
            LEFT JOIN tenant_agents ta ON ta.tenant_id = t.id AND ta.is_active = TRUE
            WHERE t.id = %s
        """, (tenant_id,))
        sp_row = cur.fetchone() or {}
        tenant_system_prompt = (sp_row.get("system_prompt") or "").strip()
    except Exception:
        tenant_system_prompt = ""

    # Plan data for assignment panel
    cur.execute("SELECT id, slug, name FROM plans WHERE is_active=TRUE ORDER BY sort_order")
    all_plans = cur.fetchall() or []

    # Messages used this billing period
    from datetime import date as _d
    period_start = customer.get("plan_period_start") or _d.today().replace(day=1)
    try:
        cur.execute("""
            SELECT COUNT(*) AS used FROM usage_events
            WHERE tenant_id=%s AND created_at >= %s
        """, (tenant_id, period_start))
        msgs_used = int((cur.fetchone() or {}).get("used") or 0)
    except Exception:
        msgs_used = 0

    cur.close(); conn.close()

    customer["balance_credits"] = tokens_to_credits(int(customer.get("token_balance") or 0))
    fn = (customer.get("first_name") or "").strip()
    ln = (customer.get("last_name")  or "").strip()
    customer["full_name"] = f"{fn} {ln}".strip() or "—"

    for inv in invs:
        inv["total_fmt"] = money_fmt(
            int(inv.get("amount_pence") or 0) + int(inv.get("vat_pence") or 0),
            inv.get("currency") or "gbp")

    return render_template("portal/admin_customer_detail.html",
                           customer=customer, keys=keys,
                           invoices=invs, audit=audit,
                           tenant_features=tenant_features,
                           doc_counts=doc_counts,
                           all_plans=all_plans,
                           msgs_used=msgs_used,
                           onboarding_other=onboarding_other,
                           tenant_system_prompt=tenant_system_prompt,
                           admin_new_plain_key=session.pop("admin_new_plain_key", None))


@portal_admin_bp.route("/customers/<int:customer_id>/system-prompt", methods=["POST"])
def customer_system_prompt_save(customer_id: int):
    r = _require_admin()
    if r: return r

    new_prompt = (request.form.get("system_prompt") or "").strip()

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(row["tenant_id"])

    # Write to active agent if one exists, otherwise fall back to tenant row
    cur.execute("""
        UPDATE tenant_agents SET system_prompt=%s, updated_at=NOW()
        WHERE tenant_id=%s AND is_active=TRUE
    """, (new_prompt, tenant_id))
    if cur.rowcount == 0:
        cur.execute("UPDATE tenants SET system_prompt=%s WHERE id=%s",
                    (new_prompt, tenant_id))
    conn.commit()
    cur.close(); conn.close()

    insert_audit_log(
        admin_username=session.get("portal_admin_username", "admin"),
        action="admin_update_system_prompt",
        tenant_id=tenant_id,
        details={"customer_id": customer_id},
    )

    flash("System prompt updated successfully.", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id) + "#system-prompt")


@portal_admin_bp.route("/customers/<int:customer_id>/credit-adjust", methods=["POST"])
def customer_credit_adjust(customer_id: int):
    r = _require_admin()
    if r: return r

    delta_credits = float(request.form.get("delta_credits") or 0)
    reason        = (request.form.get("reason") or "").strip()

    if delta_credits == 0:
        flash("Enter a non-zero credit amount.", "warning")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id   = int(row["tenant_id"])
    delta_tokens = int(delta_credits * 5000)

    cur2 = conn.cursor()
    cur2.execute("INSERT INTO tenant_balances (tenant_id, token_balance) VALUES (%s, 0) ON CONFLICT (tenant_id) DO NOTHING", (tenant_id,))
    cur2.execute("""
        UPDATE tenant_balances
        SET token_balance = GREATEST(0, token_balance + %s)
        WHERE tenant_id=%s""", (delta_tokens, tenant_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_credit_adjust",
        tenant_id=tenant_id,
        details={"delta_credits": delta_credits, "delta_tokens": delta_tokens, "reason": reason},
    )

    direction = "Added" if delta_credits > 0 else "Deducted"
    flash(f"{direction} {abs(delta_credits):.0f} credits.", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


@portal_admin_bp.route("/customers/<int:customer_id>/toggle-active", methods=["POST"])
def customer_toggle_active(customer_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT is_active FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    new_val = 0 if int(row.get("is_active") or 0) else 1
    cur2 = conn.cursor()
    cur2.execute("UPDATE customers SET is_active=%s WHERE id=%s", (new_val, customer_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(admin_username=_admin_user(),
                     action="admin_toggle_customer",
                     details={"customer_id": customer_id, "new_is_active": new_val})
    flash(f"Customer {'activated' if new_val else 'disabled'}.", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE MANAGEMENT (NEW)
# Sets which plugin features are active for a specific customer's tenant.
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/customers/<int:customer_id>/set-features", methods=["POST"])
def customer_set_features(customer_id: int):
    r = _require_admin()
    if r: return r

    # Read the feature checkboxes from the form
    feat_product_rec      = request.form.get("feat_product_recommendation") == "on"
    feat_related_products = request.form.get("feat_related_products") == "on"
    feat_cart_recovery    = request.form.get("feat_cart_recovery") == "on"
    feat_verified_specs   = request.form.get("feat_verified_specs_web_lookup") == "on"
    feat_chat_archive_30d = request.form.get("feat_chat_archive_30days") == "on"
    feat_chat_archive_unl = request.form.get("feat_chat_archive_unlimited") == "on"
    feat_wa_templates     = request.form.get("feat_wa_message_templates") == "on"

    # Cart recovery sub-settings
    recovery_popup_message = (request.form.get("cart_recovery_popup_message") or "").strip()
    recovery_incentive_pct = 0
    try:
        recovery_incentive_pct = max(0, min(50, int(request.form.get("cart_recovery_incentive_pct") or 0)))
    except (ValueError, TypeError):
        recovery_incentive_pct = 0

    # Build the features dict — add more keys here as new features are added
    features = {}
    if feat_product_rec:
        features["product_recommendation"] = True
    # Related products requires product_recommendation to also be active
    if feat_product_rec and feat_related_products:
        features["related_products"] = True
    # Cart Revenue Recovery
    if feat_cart_recovery:
        features["cart_recovery"] = True
        if recovery_popup_message:
            features["cart_recovery_popup_message"] = recovery_popup_message
        if recovery_incentive_pct > 0:
            features["cart_recovery_incentive_pct"] = recovery_incentive_pct

    # Verified Specs Lookup (Web) — allows the AI backend to browse trusted sources for numeric specs
    if feat_verified_specs:
        features["verified_specs_web_lookup"] = True

    # Chat Archive — 30 Days: search, PDF export, AI summaries, 30-day window
    # Chat Archive — Unlimited: search, all exports, AI summaries, no day limit
    # Only one tier should be active at a time; if both are ticked, Unlimited wins.
    if feat_chat_archive_unl:
        features["chat_archive_unlimited"] = True
    elif feat_chat_archive_30d:
        features["chat_archive_30days"] = True

    # WhatsApp Message Templates — cart abandonment & order status via Meta-approved templates
    if feat_wa_templates:
        features["whatsapp_message_templates"] = True

    features_json = _json.dumps(features) if features else None

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(row["tenant_id"])
    cur2 = conn.cursor()
    cur2.execute(
        "UPDATE tenants SET features=%s WHERE id=%s",
        (features_json, tenant_id)
    )
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_set_tenant_features",
        tenant_id=tenant_id,
        details={"features": features, "customer_id": customer_id},
    )

    flash("Feature settings saved. Changes take effect on the customer's next chat message.", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# TRIAL MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/customers/<int:customer_id>/trial-adjust", methods=["POST"])
def customer_trial_adjust(customer_id: int):
    r = _require_admin()
    if r: return r

    try:
        trial_days = max(1, min(3650, int(request.form.get("trial_days") or 0)))
    except (ValueError, TypeError):
        flash("Invalid number of days.", "danger")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(row["tenant_id"])

    cur.execute("""
        SELECT id, trial_activated_at FROM api_keys
        WHERE tenant_id=%s AND key_type='trial'
        ORDER BY created_at ASC LIMIT 1""", (tenant_id,))
    key = cur.fetchone()
    if not key:
        cur.close(); conn.close()
        flash("No trial key found for this customer.", "warning")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    from datetime import timedelta
    activated = key.get("trial_activated_at")
    if not activated:
        from datetime import datetime as _dt
        activated = _dt.utcnow()

    new_expiry = activated + timedelta(days=trial_days)

    cur2 = conn.cursor()
    cur2.execute(
        "UPDATE api_keys SET trial_expires_at=%s, is_active=TRUE WHERE id=%s",
        (new_expiry, key["id"])
    )
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_trial_adjusted",
        tenant_id=tenant_id,
        details={"trial_days": trial_days, "new_expiry": str(new_expiry), "customer_id": customer_id},
    )

    flash(f"Trial extended to {trial_days} days from activation. Expires {new_expiry.strftime('%Y-%m-%d')} ✅", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


@portal_admin_bp.route("/customers/<int:customer_id>/trial-reset", methods=["POST"])
def customer_trial_reset(customer_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(row["tenant_id"])

    default_days = _get_trial_default_days()

    cur.execute("""
        SELECT id FROM api_keys
        WHERE tenant_id=%s AND key_type='trial'
        ORDER BY created_at ASC LIMIT 1""", (tenant_id,))
    key = cur.fetchone()
    if not key:
        cur.close(); conn.close()
        flash("No trial key found for this customer.", "warning")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    from datetime import datetime as _dt, timedelta
    now = _dt.utcnow()
    new_expiry = now + timedelta(days=default_days)

    cur2 = conn.cursor()
    cur2.execute(
        "UPDATE api_keys SET trial_activated_at=%s, trial_expires_at=%s, is_active=TRUE, tokens_used=0 WHERE id=%s",
        (now, new_expiry, key["id"])
    )
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_trial_reset",
        tenant_id=tenant_id,
        details={"default_days": default_days, "new_expiry": str(new_expiry), "customer_id": customer_id},
    )

    flash(f"Trial reset from today. New expiry: {new_expiry.strftime('%Y-%m-%d')} ({default_days} days) ✅", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# IMPERSONATION
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/impersonate/<int:customer_id>")
def impersonate(customer_id: int):
    r = _require_admin()
    if r: return r
    session["portal_logged_in"]         = True
    session["impersonate_customer_id"]  = int(customer_id)
    insert_audit_log(admin_username=_admin_user(), action="impersonate_start",
                     details={"customer_id": customer_id})
    flash("Impersonating customer — you see the portal as they do.", "warning")
    return redirect(url_for("portal.dashboard"))


@portal_admin_bp.route("/stop-impersonate")
def stop_impersonate():
    r = _require_admin()
    if r: return r
    session.pop("impersonate_customer_id", None)
    flash("Impersonation stopped.", "success")
    return redirect(url_for("portal_admin.customers"))


# ══════════════════════════════════════════════════════════════════════════════
# API KEY REGISTRY
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/api-keys")
def api_keys():
    r = _require_admin()
    if r: return r

    q = (request.args.get("q") or "").strip().lower()

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Use ak.api_key_plain (live value from api_keys) so admin always sees
    # the key that actually authenticates, not the audit-log snapshot which
    # can diverge if a key was regenerated after creation.
    if q:
        cur.execute("""
            SELECT al.id,
                   COALESCE(ak.api_key_plain, al.api_key_plain) AS api_key_plain,
                   al.api_key_last4, al.website, al.key_type,
                   al.tenant_id, al.api_key_id, al.created_at, al.admin_username, al.details,
                   t.name AS tenant_name, t.domain,
                   ak.is_active, ak.tokens_used, ak.token_limit,
                   ak.trial_expires_at
            FROM audit_logs al
            JOIN tenants t ON t.id = al.tenant_id
            LEFT JOIN api_keys ak ON ak.id = al.api_key_id
            WHERE al.action='create_key'
              AND (LOWER(al.website) LIKE %s
                OR LOWER(t.name) LIKE %s
                OR LOWER(al.api_key_last4) LIKE %s)
            ORDER BY al.created_at DESC
            LIMIT 500""",
            (f"%{q}%", f"%{q}%", f"%{q}%"))
    else:
        cur.execute("""
            SELECT al.id,
                   COALESCE(ak.api_key_plain, al.api_key_plain) AS api_key_plain,
                   al.api_key_last4, al.website, al.key_type,
                   al.tenant_id, al.api_key_id, al.created_at, al.admin_username, al.details,
                   t.name AS tenant_name, t.domain,
                   ak.is_active, ak.tokens_used, ak.token_limit,
                   ak.trial_expires_at
            FROM audit_logs al
            JOIN tenants t ON t.id = al.tenant_id
            LEFT JOIN api_keys ak ON ak.id = al.api_key_id
            WHERE al.action='create_key'
            ORDER BY al.created_at DESC
            LIMIT 500""")

    rows = cur.fetchall() or []
    cur.close(); conn.close()

    for row in rows:
        row["credits_used"] = tokens_to_credits(int(row.get("tokens_used") or 0))
        # Status
        is_active = row.get("is_active")
        if is_active is None:
            row["status"] = "Unknown"
        elif int(is_active) == 0:
            row["status"] = "Revoked"
        else:
            row["status"] = "Active"

    return render_template("portal/admin_api_keys.html", keys=rows, q=q)


@portal_admin_bp.route("/api-keys/<int:key_id>/revoke", methods=["POST"])
def api_keys_revoke(key_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, tenant_id, website, key_type FROM api_keys WHERE id=%s", (key_id,))
    k = cur.fetchone()
    if k:
        cur2 = conn.cursor()
        cur2.execute("UPDATE api_keys SET is_active=FALSE WHERE id=%s", (key_id,))
        conn.commit()
        cur2.close()
        insert_audit_log(admin_username=_admin_user(), action="admin_revoke_key",
                         tenant_id=k.get("tenant_id"), website=k.get("website"),
                         key_type=k.get("key_type"), api_key_id=key_id)
    cur.close(); conn.close()
    flash("Key revoked.", "success")
    return redirect(url_for("portal_admin.api_keys"))


@portal_admin_bp.route("/api-keys/<int:key_id>/reactivate", methods=["POST"])
def api_keys_reactivate(key_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, tenant_id, website, key_type FROM api_keys WHERE id=%s", (key_id,))
    k = cur.fetchone()
    if k:
        cur2 = conn.cursor()
        if k.get("key_type") == "trial":
            default_days = _get_trial_default_days()
            from datetime import datetime as _dt, timedelta
            new_expiry = _dt.utcnow() + timedelta(days=default_days)
            cur2.execute(
                "UPDATE api_keys SET is_active=TRUE, trial_expires_at=%s WHERE id=%s",
                (new_expiry, key_id)
            )
        else:
            cur2.execute("UPDATE api_keys SET is_active=TRUE WHERE id=%s", (key_id,))
        conn.commit()
        cur2.close()
        insert_audit_log(admin_username=_admin_user(), action="admin_reactivate_key",
                         tenant_id=k.get("tenant_id"), website=k.get("website"),
                         key_type=k.get("key_type"), api_key_id=key_id)
    cur.close(); conn.close()
    flash("Key reactivated.", "success")
    return redirect(url_for("portal_admin.api_keys"))


def _admin_generate_api_key_and_hash(length: int = 28):
    """Same algorithm as portal_routes.py — keep in sync."""
    alphabet = string.ascii_letters + string.digits
    plain_key = ''.join(secrets.choice(alphabet) for _ in range(length))
    hashed_key = bcrypt.hashpw(plain_key.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    return plain_key, hashed_key


@portal_admin_bp.route("/customers/<int:customer_id>/api-keys/create", methods=["POST"])
def customer_api_key_create(customer_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Look up customer → tenant
    cur.execute("""
        SELECT c.id, c.email, t.id AS tenant_id, t.domain
        FROM customers c
        JOIN tenants t ON t.id = c.tenant_id
        WHERE c.id=%s""", (customer_id,))
    customer = cur.fetchone()
    if not customer:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(customer["tenant_id"])
    domain    = customer.get("domain") or ""

    # Safety: refuse if an active paid key already exists for this tenant
    cur.execute("""
        SELECT id FROM api_keys
        WHERE tenant_id=%s AND key_type='paid' AND is_active=TRUE
        LIMIT 1""", (tenant_id,))
    if cur.fetchone():
        cur.close(); conn.close()
        flash("An active paid key already exists for this tenant. Revoke it first.", "warning")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    plain_key, hashed_key = _admin_generate_api_key_and_hash()
    last4 = plain_key[-4:]

    try:
        cur2 = conn.cursor()
        cur2.execute("""
            INSERT INTO api_keys
                (tenant_id, api_key_hash, api_key_plain, is_active, website, key_type, token_limit, tokens_used)
            VALUES (%s, %s, %s, TRUE, %s, 'paid', NULL, 0)
            RETURNING id""",
            (tenant_id, hashed_key, plain_key, domain))
        api_key_id = cur2.fetchone()[0]
        conn.commit()
        cur2.close()
    except Exception as e:
        conn.rollback()
        cur.close(); conn.close()
        flash(f"Could not create key: {e}", "danger")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="create_key",
        tenant_id=tenant_id,
        website=domain,
        key_type="paid",
        api_key_id=api_key_id,
        api_key_last4=last4,
        api_key_plain=plain_key,
        details={"created_from": "admin_portal"},
    )

    # Store plain key in session — shown once on the customer detail page
    session["admin_new_plain_key"] = plain_key
    flash("API key created. Copy it now — it will not be shown again.", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# CREDIT PACKAGES (admin)
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/credit-packages", methods=["GET", "POST"])
def credit_packages():
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        name        = (request.form.get("name")        or "").strip()
        credits     = int(request.form.get("credits")  or 0)
        price_pence = int(float(request.form.get("price_gbp") or 0) * 100)
        vat_rate    = float(request.form.get("vat_rate")   or 20.0)
        is_active = 1 if request.form.get("is_active") == "on" else 0
        sort_order  = int(request.form.get("sort_order")   or 0)

        # Stage 3 — package type and billing period
        # package_type : 'topup' (default, existing behaviour) or 'subscription'
        # billing_period: 'monthly' or 'annual' — only used for subscriptions
        raw_pkg_type    = (request.form.get("package_type") or "topup").strip()
        package_type    = raw_pkg_type if raw_pkg_type in ("topup", "subscription") else "topup"
        raw_billing     = (request.form.get("billing_period") or "").strip()
        billing_period  = raw_billing if raw_billing in ("monthly", "annual") else None
        # billing_period is only meaningful for subscription packages
        if package_type == "topup":
            billing_period = None

        # Read feature checkboxes
        feat_product_rec      = request.form.get("feature_product_recommendation") == "on"
        feat_related_products = request.form.get("feature_related_products") == "on"
        feat_cart_recovery    = request.form.get("feature_cart_recovery") == "on"
        feat_verified_specs  = request.form.get("feature_verified_specs_web_lookup") == "on"
        feat_chat_archive_30d = request.form.get("feature_chat_archive_30days") == "on"
        feat_chat_archive_unl = request.form.get("feature_chat_archive_unlimited") == "on"
        feat_wa_templates     = request.form.get("feature_wa_message_templates") == "on"
        features_dict = {}
        if feat_product_rec:
            features_dict["product_recommendation"] = True
        # Related products requires product_recommendation to also be active
        if feat_product_rec and feat_related_products:
            features_dict["related_products"] = True
        # Intelligent Cart Revenue Recovery
        if feat_cart_recovery:
            features_dict["cart_recovery"] = True
        # Verified Specs Lookup (Web)
        if feat_verified_specs:
            features_dict["verified_specs_web_lookup"] = True
        # Chat Archive tiers — only one can be set; Unlimited wins if both ticked
        if feat_chat_archive_unl:
            features_dict["chat_archive_unlimited"] = True
        elif feat_chat_archive_30d:
            features_dict["chat_archive_30days"] = True
        # WhatsApp Message Templates
        if feat_wa_templates:
            features_dict["whatsapp_message_templates"] = True
        # Custom features — one per line entered by admin
        custom_text = (request.form.get("custom_features_text") or "").strip()
        custom_list = [line.strip() for line in custom_text.splitlines() if line.strip()]
        if custom_list:
            features_dict["custom_features"] = custom_list
        features_json = _json.dumps(features_dict) if features_dict else None

        if not name or credits <= 0 or price_pence <= 0:
            flash("Name, credits and price are required.", "danger")
        else:
            cur2 = conn.cursor()
            cur2.execute("""
                INSERT INTO credit_packages (name, credits, price_pence, vat_rate, is_active, sort_order, features, package_type, billing_period)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (name, credits, price_pence, vat_rate, is_active, sort_order, features_json, package_type, billing_period))
            conn.commit()
            cur2.close()
            flash("Package added.", "success")
        return redirect(url_for("portal_admin.credit_packages"))

    cur.execute("SELECT * FROM credit_packages ORDER BY sort_order ASC, id ASC")
    packages = cur.fetchall() or []
    cur.close(); conn.close()

    for p in packages:
        p["price_fmt"] = money_fmt(int(p.get("price_pence") or 0), p.get("currency") or "gbp")
        # Parse the features JSON for the template
        raw_feat = p.get("features")
        if raw_feat:
            try:
                p["features_parsed"] = _json.loads(raw_feat) if isinstance(raw_feat, str) else raw_feat
            except Exception:
                p["features_parsed"] = {}
        else:
            p["features_parsed"] = {}

    return render_template("portal/admin_packages.html", packages=packages)


@portal_admin_bp.route("/credit-packages/<int:pkg_id>/toggle")
def credit_packages_toggle(pkg_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT is_active FROM credit_packages WHERE id=%s", (pkg_id,))
    row = cur.fetchone() or {}
    new_val = 0 if int(row.get("is_active") or 0) else 1
    cur2 = conn.cursor()
    cur2.execute("UPDATE credit_packages SET is_active=%s WHERE id=%s", (new_val, pkg_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    flash("Package updated.", "success")
    return redirect(url_for("portal_admin.credit_packages"))


@portal_admin_bp.route("/credit-packages/<int:pkg_id>/delete", methods=["POST"])
def credit_packages_delete(pkg_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT name FROM credit_packages WHERE id=%s", (pkg_id,))
    row = cur.fetchone()
    if row:
        cur2 = conn.cursor()
        cur2.execute("DELETE FROM credit_packages WHERE id=%s", (pkg_id,))
        conn.commit()
        cur2.close()
        insert_audit_log(
            admin_username=_admin_user(),
            action="admin_delete_package",
            details={"package_id": pkg_id, "package_name": row.get("name")},
        )
        flash(f"Package '{row.get('name')}' deleted.", "success")
    else:
        flash("Package not found.", "danger")
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.credit_packages"))


@portal_admin_bp.route("/credit-packages/<int:pkg_id>/edit", methods=["POST"])
def credit_packages_edit(pkg_id: int):
    r = _require_admin()
    if r: return r

    name        = (request.form.get("name")        or "").strip()
    credits     = int(request.form.get("credits")  or 0)
    price_pence = int(float(request.form.get("price_gbp") or 0) * 100)
    vat_rate    = float(request.form.get("vat_rate")   or 20.0)
    is_active=TRUE if request.form.get("is_active") == "on" else 0
    sort_order  = int(request.form.get("sort_order")   or 0)

    # Stage 3 — package type and billing period
    raw_pkg_type_e   = (request.form.get("package_type") or "topup").strip()
    package_type_e   = raw_pkg_type_e if raw_pkg_type_e in ("topup", "subscription") else "topup"
    raw_billing_e    = (request.form.get("billing_period") or "").strip()
    billing_period_e = raw_billing_e if raw_billing_e in ("monthly", "annual") else None
    if package_type_e == "topup":
        billing_period_e = None

    feat_product_rec      = request.form.get("feature_product_recommendation") == "on"
    feat_related_products = request.form.get("feature_related_products") == "on"
    feat_cart_recovery    = request.form.get("feature_cart_recovery") == "on"
    feat_verified_specs   = request.form.get("feature_verified_specs_web_lookup") == "on"
    feat_chat_archive_30d = request.form.get("feature_chat_archive_30days") == "on"
    feat_chat_archive_unl = request.form.get("feature_chat_archive_unlimited") == "on"
    feat_wa_templates     = request.form.get("feature_wa_message_templates") == "on"
    features_dict = {}
    if feat_product_rec:
        features_dict["product_recommendation"] = True
    if feat_product_rec and feat_related_products:
        features_dict["related_products"] = True
    if feat_cart_recovery:
        features_dict["cart_recovery"] = True
    if feat_verified_specs:
        features_dict["verified_specs_web_lookup"] = True
    # Chat Archive tiers — only one can be set; Unlimited wins if both ticked
    if feat_chat_archive_unl:
        features_dict["chat_archive_unlimited"] = True
    elif feat_chat_archive_30d:
        features_dict["chat_archive_30days"] = True
    # WhatsApp Message Templates
    if feat_wa_templates:
        features_dict["whatsapp_message_templates"] = True
    # Custom features — one per line entered by admin
    custom_text = (request.form.get("custom_features_text") or "").strip()
    custom_list = [line.strip() for line in custom_text.splitlines() if line.strip()]
    if custom_list:
        features_dict["custom_features"] = custom_list
    features_json = _json.dumps(features_dict) if features_dict else None

    if not name or credits <= 0 or price_pence <= 0:
        flash("Name, credits and price are required.", "danger")
        return redirect(url_for("portal_admin.credit_packages"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE credit_packages
        SET name=%s, credits=%s, price_pence=%s, vat_rate=%s,
            is_active=%s, sort_order=%s, features=%s,
            package_type=%s, billing_period=%s
        WHERE id=%s""",
        (name, credits, price_pence, vat_rate, is_active, sort_order, features_json,
         package_type_e, billing_period_e, pkg_id))
    conn.commit()
    cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_edit_package",
        details={"package_id": pkg_id, "name": name, "credits": credits,
                 "price_pence": price_pence, "features": features_dict,
                 "package_type": package_type_e, "billing_period": billing_period_e},
    )
    flash(f"Package '{name}' updated.", "success")
    return redirect(url_for("portal_admin.credit_packages"))


# ══════════════════════════════════════════════════════════════════════════════
# PLUGIN DOWNLOADS (admin upload / manage)
# Customers download plugins from the onboarding wizard.
# Admin uploads zip files here and can replace them when new versions are ready.
# ══════════════════════════════════════════════════════════════════════════════

import os as _os
import werkzeug.utils as _wu

PLUGIN_UPLOAD_DIR = "/root/api-key-manager/static/plugin_zips"

@portal_admin_bp.route("/plugins", methods=["GET", "POST"])
def plugins():
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "upload":
            plugin_key   = (request.form.get("plugin_key")   or "").strip().lower()
            display_name = (request.form.get("display_name") or "").strip()
            version      = (request.form.get("version")      or "").strip()
            f = request.files.get("plugin_file")

            if not plugin_key or not display_name or not f or not f.filename:
                flash("Plugin key, display name, and zip file are all required.", "danger")
            elif not f.filename.lower().endswith(".zip"):
                flash("Only .zip files are allowed.", "danger")
            else:
                _os.makedirs(PLUGIN_UPLOAD_DIR, exist_ok=True)
                safe_name = _wu.secure_filename(f.filename)
                dest_path = _os.path.join(PLUGIN_UPLOAD_DIR, safe_name)
                f.save(dest_path)

                cur2 = conn.cursor()
                cur2.execute("""
                    INSERT INTO plugin_downloads (plugin_key, display_name, filename, file_path, version)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (plugin_key) DO UPDATE SET
                        display_name=EXCLUDED.display_name, filename=EXCLUDED.filename,
                        file_path=EXCLUDED.file_path, version=EXCLUDED.version,
                        uploaded_at=CURRENT_TIMESTAMP""",
                    (plugin_key, display_name, safe_name, dest_path, version or None))
                conn.commit()
                cur2.close()

                insert_audit_log(admin_username=_admin_user(), action="plugin_upload",
                                 details={"plugin_key": plugin_key, "filename": safe_name, "version": version})
                flash(f"Plugin '{display_name}' uploaded successfully.", "success")

        elif action == "delete":
            plugin_key = (request.form.get("plugin_key") or "").strip()
            cur.execute("SELECT * FROM plugin_downloads WHERE plugin_key=%s", (plugin_key,))
            row = cur.fetchone()
            if row:
                # Remove file from disk
                try:
                    if _os.path.exists(row["file_path"]):
                        _os.remove(row["file_path"])
                except Exception:
                    pass
                cur2 = conn.cursor()
                cur2.execute("DELETE FROM plugin_downloads WHERE plugin_key=%s", (plugin_key,))
                conn.commit()
                cur2.close()
                insert_audit_log(admin_username=_admin_user(), action="plugin_delete",
                                 details={"plugin_key": plugin_key})
                flash("Plugin removed.", "success")

        return redirect(url_for("portal_admin.plugins"))

    cur.execute("SELECT * FROM plugin_downloads ORDER BY uploaded_at DESC")
    plugins_list = cur.fetchall() or []
    cur.close(); conn.close()

    return render_template("portal/admin_plugins.html", plugins=plugins_list)


# ══════════════════════════════════════════════════════════════════════════════
# INVOICES (admin)
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/invoices")
def invoices():
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT i.id, i.invoice_number, i.credits, i.amount_pence, i.vat_pence,
               i.currency, i.status, i.created_at,
               c.email AS customer_email,
               CONCAT(COALESCE(c.first_name,''),' ',COALESCE(c.last_name,'')) AS customer_name,
               t.name AS tenant_name, t.domain
        FROM invoices i
        JOIN customers c ON c.id=i.customer_id
        JOIN tenants   t ON t.id=i.tenant_id
        ORDER BY i.created_at DESC LIMIT 500""")
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    for row in rows:
        total = int(row.get("amount_pence") or 0) + int(row.get("vat_pence") or 0)
        row["total_fmt"] = money_fmt(total, row.get("currency") or "gbp")

    return render_template("portal/admin_invoices.html", invoices=rows)


@portal_admin_bp.route("/invoice/<int:invoice_id>/download")
def invoice_download(invoice_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM invoices WHERE id=%s", (invoice_id,))
    inv = cur.fetchone()
    cur.close(); conn.close()

    if not inv or inv.get("status") != "paid" or not inv.get("pdf_path"):
        flash("Invoice PDF not available.", "warning")
        return redirect(url_for("portal_admin.invoices"))

    if not os.path.exists(inv["pdf_path"]):
        flash("File missing.", "danger")
        return redirect(url_for("portal_admin.invoices"))

    return send_file(inv["pdf_path"], as_attachment=True,
                     download_name=f"{inv['invoice_number']}.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# CART REVENUE RECOVERY DASHBOARD (admin)
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/recovery-queue")
def recovery_queue():
    r = _require_admin()
    if r: return r

    # Optional filters from query string
    status_filter  = (request.args.get("status")    or "").strip() or None
    tenant_filter  = (request.args.get("tenant_id") or "").strip()
    tenant_id_filter = int(tenant_filter) if tenant_filter.isdigit() else None

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # ── KPI counts ────────────────────────────────────────────────────────
        if tenant_id_filter:
            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN status='pending'     THEN 1 ELSE 0 END) AS pending,
                    SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END) AS in_progress,
                    SUM(CASE WHEN status='recovered'   THEN 1 ELSE 0 END) AS recovered,
                    SUM(CASE WHEN status='expired'     THEN 1 ELSE 0 END) AS expired
                FROM abandonment_queue
                WHERE tenant_id = %s
            """, (tenant_id_filter,))
        else:
            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN status='pending'     THEN 1 ELSE 0 END) AS pending,
                    SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END) AS in_progress,
                    SUM(CASE WHEN status='recovered'   THEN 1 ELSE 0 END) AS recovered,
                    SUM(CASE WHEN status='expired'     THEN 1 ELSE 0 END) AS expired
                FROM abandonment_queue
            """)
        stats = cur.fetchone() or {}
        for k in ("total", "pending", "in_progress", "recovered", "expired"):
            stats[k] = int(stats.get(k) or 0)

        # Conversion rate (recovered / total active sessions that reached recovery)
        eligible = stats["in_progress"] + stats["recovered"] + stats["expired"]
        stats["conversion_rate"] = (
            round(stats["recovered"] / eligible * 100, 1) if eligible > 0 else 0.0
        )

        # ── Queue rows ────────────────────────────────────────────────────────
        where_parts  = []
        params: list = []

        if tenant_id_filter:
            where_parts.append("q.tenant_id = %s")
            params.append(tenant_id_filter)

        if status_filter and status_filter in ("pending", "in_progress", "recovered", "expired"):
            where_parts.append("q.status = %s")
            params.append(status_filter)

        where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
        params.append(200)  # LIMIT

        cur.execute(f"""
            SELECT q.id, q.tenant_id, q.session_id, q.intent_score, q.priority,
                   q.cart_value, q.customer_email, q.status, q.touches_sent,
                   q.expires_at, q.created_at, q.updated_at,
                   t.name AS tenant_name, t.domain
            FROM abandonment_queue q
            LEFT JOIN tenants t ON t.id = q.tenant_id
            {where_clause}
            ORDER BY q.updated_at DESC
            LIMIT %s
        """, params)
        rows = cur.fetchall() or []

        # ── Tenant list for filter dropdown ───────────────────────────────────
        cur.execute("""
            SELECT DISTINCT t.id, t.name, t.domain
            FROM tenants t
            INNER JOIN abandonment_queue q ON q.tenant_id = t.id
            ORDER BY t.name ASC LIMIT 200
        """)
        tenants_list = cur.fetchall() or []

    finally:
        try: cur.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass

    # Format cart values for display
    for row in rows:
        cv = row.get("cart_value")
        row["cart_value_fmt"] = f"£{float(cv):.2f}" if cv else "—"

    return render_template(
        "portal/admin_recovery_queue.html",
        stats=stats,
        rows=rows,
        tenants_list=tenants_list,
        status_filter=status_filter or "",
        tenant_id_filter=tenant_id_filter or "",
    )


# ══════════════════════════════════════════════════════════════════════════════
# ADMIN SETTINGS — Password Reset & Customer Email Change
# ══════════════════════════════════════════════════════════════════════════════

import bcrypt as _bcrypt_admin


def _verify_admin_password(plain: str, stored_admin: dict) -> bool:
    """
    Supports both legacy plain-text passwords and new bcrypt hashes.
    If password_hash column exists and is set, use bcrypt.
    Otherwise fall back to plain-text comparison (legacy).
    """
    pw_hash = stored_admin.get("password_hash")
    if pw_hash:
        try:
            return _bcrypt_admin.checkpw(plain.encode("utf-8"), pw_hash.encode("utf-8"))
        except Exception:
            return False
    # Legacy plain-text fallback
    return plain == stored_admin.get("password", "")


def _hash_admin_password(plain: str) -> str:
    return _bcrypt_admin.hashpw(plain.encode("utf-8"), _bcrypt_admin.gensalt()).decode("utf-8")


def _get_trial_default_days() -> int:
    """Read trial_default_days from portal_settings table.
    Falls back to 14 if the table doesn't exist yet or the key is missing."""
    try:
        conn = get_db_connection()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT setting_value FROM portal_settings WHERE setting_key='trial_default_days'")
        row = cur.fetchone()
        cur.close(); conn.close()
        if row and row.get("setting_value"):
            return int(row["setting_value"])
    except Exception:
        pass
    return 14  # default


@portal_admin_bp.route("/settings", methods=["GET"])
def admin_settings():
    r = _require_admin()
    if r: return r
    username          = _admin_user()
    trial_default_days = _get_trial_default_days()
    return render_template("portal/admin_settings.html",
                           username=username,
                           trial_default_days=trial_default_days)


@portal_admin_bp.route("/onboarding-qr")
def onboarding_qr():
    """Serve the WhatsApp onboarding QR code card as a downloadable PNG."""
    r = _require_admin()
    if r: return r
    import os
    qr_path = os.path.join(
        os.path.dirname(__file__), "static", "portal", "whatsapp_setup_qr_card.png"
    )
    return send_file(qr_path, mimetype="image/png",
                     as_attachment=True, download_name="phixtra_whatsapp_setup_qr.png")


@portal_admin_bp.route("/settings/trial-days", methods=["POST"])
def admin_save_trial_days():
    """Save the default trial duration (days) into portal_settings."""
    r = _require_admin()
    if r: return r

    raw = (request.form.get("trial_default_days") or "").strip()
    try:
        days = int(raw)
        if not (1 <= days <= 365):
            raise ValueError
    except ValueError:
        flash("Please enter a number between 1 and 365.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    try:
        conn = get_db_connection()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO portal_settings (setting_key, setting_value)
            VALUES ('trial_default_days', %s)
            ON CONFLICT (setting_key) DO UPDATE SET setting_value = EXCLUDED.setting_value
        """, (str(days),))
        conn.commit()
        cur.close(); conn.close()
        insert_audit_log(
            admin_username=_admin_user(),
            action="admin_trial_days_updated",
            details={"trial_default_days": days},
        )
        flash(f"Default trial duration updated to {days} days ✅", "success")
    except Exception as e:
        print("⚠️ admin_save_trial_days error:", e)
        flash("Could not save setting. Please try again.", "danger")

    return redirect(url_for("portal_admin.admin_settings"))


@portal_admin_bp.route("/settings/password", methods=["POST"])
def admin_change_password():
    """Allow an admin to change their own login password."""
    r = _require_admin()
    if r: return r

    username    = _admin_user()
    current_pw  = (request.form.get("current_password") or "").strip()
    new_pw      = (request.form.get("new_password")     or "").strip()
    confirm_pw  = (request.form.get("confirm_password") or "").strip()

    if not current_pw or not new_pw or not confirm_pw:
        flash("All password fields are required.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    if new_pw != confirm_pw:
        flash("New passwords do not match.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    if len(new_pw) < 8:
        flash("New password must be at least 8 characters.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM admin_users WHERE username=%s", (username,))
    admin = cur.fetchone()

    if not admin:
        cur.close(); conn.close()
        flash("Admin account not found.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    if not _verify_admin_password(current_pw, admin):
        cur.close(); conn.close()
        flash("Current password is incorrect.", "danger")
        return redirect(url_for("portal_admin.admin_settings"))

    new_hash = _hash_admin_password(new_pw)
    cur2 = conn.cursor()
    # Update both password (legacy) and password_hash (bcrypt)
    cur2.execute(
        "UPDATE admin_users SET password=%s, password_hash=%s WHERE username=%s",
        (new_pw, new_hash, username)
    )
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(action="admin_password_changed",
                     admin_username=username,
                     details={"changed_by": username})
    flash("Admin password changed successfully ✅", "success")
    return redirect(url_for("portal_admin.admin_settings"))


@portal_admin_bp.route("/customers/<int:customer_id>/change-email", methods=["POST"])
def customer_change_email(customer_id: int):
    """Admin can update a customer's email address."""
    r = _require_admin()
    if r: return r

    new_email = (request.form.get("new_email") or "").strip().lower()
    if not new_email or "@" not in new_email or "." not in new_email.split("@")[-1]:
        flash("Please enter a valid email address.", "danger")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Check the customer exists
    cur.execute("SELECT id, email, tenant_id FROM customers WHERE id=%s", (customer_id,))
    c = cur.fetchone()
    if not c:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    old_email = c.get("email") or ""

    # Check for duplicate email
    cur.execute("SELECT id FROM customers WHERE email=%s AND id != %s", (new_email, customer_id))
    dup = cur.fetchone()
    if dup:
        cur.close(); conn.close()
        flash(f"Email address {new_email} is already in use by another account.", "danger")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    cur2 = conn.cursor()
    cur2.execute("UPDATE customers SET email=%s WHERE id=%s", (new_email, customer_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        action="admin_customer_email_changed",
        admin_username=_admin_user(),
        tenant_id=int(c.get("tenant_id") or 0),
        details={"customer_id": customer_id, "old_email": old_email, "new_email": new_email},
    )
    flash(f"Customer email updated from {old_email} → {new_email} ✅", "success")
    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


@portal_admin_bp.route("/customers/<int:customer_id>/send-reset", methods=["POST"])
def customer_send_reset(customer_id: int):
    """Admin triggers a password reset email for a customer."""
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, first_name, email FROM customers WHERE id=%s", (customer_id,))
    c = cur.fetchone()

    if not c:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    from portal_utils import make_token, utc_now_naive, send_email
    from datetime import timedelta

    token   = make_token(24)
    expires = utc_now_naive() + timedelta(hours=2)

    cur2 = conn.cursor()
    cur2.execute("UPDATE customers SET reset_token=%s, reset_expires_at=%s WHERE id=%s",
                 (token, expires, customer_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    PORTAL_BASE_URL = os.getenv("PORTAL_BASE_URL", "https://portal.phixtra.com").rstrip("/")
    link    = f"{PORTAL_BASE_URL}/reset?token={token}"
    greeting = (c.get("first_name") or "there").strip()
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:520px">
      <h2 style="color:#030C18">Reset your password</h2>
      <p>Hi {greeting},</p>
      <p>A PhiXtra admin has sent you a password reset link.</p>
      <p>Click below to set a new password. This link expires in 2 hours.</p>
      <p><a href="{link}" style="background:#030C18;color:#fff;padding:10px 18px;border-radius:12px;
                                  text-decoration:none;display:inline-block">Reset password</a></p>
      <p style="color:#888;font-size:12px">If you didn't request this, please ignore this email.</p>
    </div>"""
    sent = send_email(c["email"], "Reset your PhiXtra password", html, text_body=f"Reset: {link}")

    insert_audit_log(
        action="admin_sent_customer_password_reset",
        admin_username=_admin_user(),
        details={"customer_id": customer_id, "email": c["email"], "email_sent": sent},
    )

    if sent:
        flash(f"Password reset email sent to {c['email']} ✅", "success")
    else:
        flash(f"Reset token created but email could not be sent. Link: {link}", "warning")

    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# SEARCH INDEX MANAGEMENT (admin-only)
# Deletes documents from the pgvector documents table for a specific tenant.
# After deleting, the customer must run Full Sync from their WordPress plugin.
# ══════════════════════════════════════════════════════════════════════════════

_VALID_DOC_TYPES = {"product", "post", "page", "order", "customer"}


@portal_admin_bp.route("/customers/<int:customer_id>/rebuild-index", methods=["POST"])
def customer_rebuild_index(customer_id: int):
    r = _require_admin()
    if r: return r

    doc_type = (request.form.get("doc_type") or "").strip().lower()

    if doc_type and doc_type not in _VALID_DOC_TYPES:
        flash(f"Invalid document type: {doc_type!r}. Must be one of: {', '.join(sorted(_VALID_DOC_TYPES))}.", "danger")
        return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT tenant_id FROM customers WHERE id=%s", (customer_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        flash("Customer not found.", "danger")
        return redirect(url_for("portal_admin.customers"))

    tenant_id = int(row["tenant_id"])

    cur2 = conn.cursor()
    if doc_type:
        cur2.execute(
            "DELETE FROM documents WHERE tenant_id = %s AND type = %s",
            (tenant_id, doc_type),
        )
    else:
        # Delete all types except verified_spec (same rule as /sync/rebuild-index endpoint)
        cur2.execute(
            "DELETE FROM documents WHERE tenant_id = %s AND type != 'verified_spec'",
            (tenant_id,),
        )

    deleted = cur2.rowcount
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    insert_audit_log(
        admin_username=_admin_user(),
        action="admin_rebuild_index",
        tenant_id=tenant_id,
        details={
            "customer_id": customer_id,
            "doc_type": doc_type or "all",
            "deleted_count": deleted,
        },
    )

    if doc_type:
        flash(
            f"Deleted {deleted:,} '{doc_type}' documents for this tenant. "
            f"Ask them to run Full Sync from WordPress → PhiXtra Export → PhiXtra Sync tab.",
            "success",
        )
    else:
        flash(
            f"Deleted {deleted:,} documents (all types) for this tenant. "
            f"Ask them to run Full Sync from WordPress → PhiXtra Export → PhiXtra Sync tab.",
            "success",
        )

    return redirect(url_for("portal_admin.customer_detail", customer_id=customer_id))


# ══════════════════════════════════════════════════════════════════════════════
# WHATSAPP MERCHANT ONBOARDING
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/onboard-wa", methods=["GET", "POST"])
def onboard_wa():
    r = _require_admin()
    if r: return r

    if request.method == "POST":
        from portal_routes import provision_whatsapp_merchant, _normalise_phone
        raw_phone     = (request.form.get("phone")         or "").strip()
        business_name = (request.form.get("business_name") or "").strip()
        notes         = (request.form.get("notes")         or "").strip()

        if not raw_phone or not business_name:
            flash("Phone number and business name are both required.", "danger")
            return redirect(url_for("portal_admin.onboard_wa"))

        phone = _normalise_phone(raw_phone)
        if not phone:
            flash(f"Could not parse phone number: {raw_phone!r}. Use +234… or 0801… format.", "danger")
            return redirect(url_for("portal_admin.onboard_wa"))

        try:
            result = provision_whatsapp_merchant(phone, business_name)
            insert_audit_log(
                action="admin_onboarded_wa_merchant",
                admin_username=_admin_user(),
                tenant_id=result["tenant_id"],
                details={"phone": phone, "business_name": business_name, "notes": notes},
            )
            flash(
                f"✅ <strong>{business_name}</strong> provisioned. "
                f"Tenant #{result['tenant_id']} · Customer #{result['customer_id']}. "
                f"They can log in at <a href='https://portal.phixtra.com/wa-login' target='_blank'>"
                f"portal.phixtra.com/wa-login</a> using <strong>{phone}</strong>.",
                "success"
            )
        except Exception as e:
            flash(f"Provisioning failed: {e}", "danger")

        return redirect(url_for("portal_admin.onboard_wa"))

    # ── GET: load existing WA merchants ──────────────────────────────────────
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            t.id            AS tenant_id,
            t.name          AS business_name,
            t.status,
            t.created_at    AS provisioned_at,
            c.phone_number,
            c.id            AS customer_id,
            c.is_active,
            COALESCE(tb.token_balance, 0) AS token_balance
        FROM tenants t
        JOIN customers c        ON c.tenant_id = t.id
        LEFT JOIN tenant_balances tb ON tb.tenant_id = t.id
        WHERE t.source_type = 'whatsapp'
        ORDER BY t.created_at DESC
        LIMIT 100
    """)
    merchants = cur.fetchall() or []
    cur.close(); conn.close()

    return render_template("portal/admin_onboard_wa.html",
                           merchants=merchants)


# ══════════════════════════════════════════════════════════════════════════════
# PHONE CATALOGUE
# ══════════════════════════════════════════════════════════════════════════════

_CAT_COLS = [
    "product_id","brand","model_name","variant_name","release_year",
    "price_category","network_type","screen_size_inches","display_type",
    "refresh_rate_hz","screen_resolution","chipset_model","ram","storage",
    "battery_capacity_mah","fast_charging_watts","rear_camera_main_mp",
    "front_camera_mp","video_recording","gaming_rating","battery_performance",
    "camera_quality_rating","wifi_version","bluetooth_version","nfc",
    "body_material","water_resistance","fingerprint_type","available_colors",
    "nigeria_market_price_naira","best_for","search_intent_tags",
    "ai_summary","ai_sales_pitch","is_active",
]

_FILTER_COLS = ["brand","price_category","network_type","display_type","nfc","is_active"]

PAGE_SIZE = 50


def _catalogue_brands(cur):
    cur.execute("SELECT DISTINCT brand FROM phone_catalogue WHERE brand IS NOT NULL ORDER BY brand")
    return [r["brand"] for r in (cur.fetchall() or [])]


@portal_admin_bp.route("/catalogue/phones")
def catalogue():
    r = _require_admin()
    if r: return r

    q       = (request.args.get("q") or "").strip()
    brand   = (request.args.get("brand") or "").strip()
    pcat    = (request.args.get("price_category") or "").strip()
    network = (request.args.get("network_type") or "").strip()
    nfc     = (request.args.get("nfc") or "").strip()
    active  = request.args.get("is_active", "")
    page    = max(1, int(request.args.get("page", 1)))

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    conditions = []
    params     = []

    if q:
        conditions.append(
            "to_tsvector('english', COALESCE(brand,'') || ' ' || COALESCE(model_name,'') || ' ' || COALESCE(variant_name,'') || ' ' || COALESCE(search_intent_tags,'')) @@ plainto_tsquery('english', %s)"
        )
        params.append(q)
    if brand:
        conditions.append("brand = %s"); params.append(brand)
    if pcat:
        conditions.append("price_category = %s"); params.append(pcat)
    if network:
        conditions.append("network_type = %s"); params.append(network)
    if nfc:
        conditions.append("nfc = %s"); params.append(nfc)
    if active in ("true", "false"):
        conditions.append("is_active = %s"); params.append(active == "true")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    cur.execute(f"SELECT COUNT(*) AS n FROM phone_catalogue {where}", params)
    total = cur.fetchone()["n"]

    offset = (page - 1) * PAGE_SIZE
    cur.execute(
        f"""SELECT id, product_id, brand, model_name, variant_name, release_year,
                   price_category, network_type, nigeria_market_price_naira, is_active
            FROM phone_catalogue {where}
            ORDER BY brand, model_name, variant_name
            LIMIT %s OFFSET %s""",
        params + [PAGE_SIZE, offset],
    )
    rows = cur.fetchall() or []

    brands   = _catalogue_brands(cur)
    cur.execute("SELECT DISTINCT price_category FROM phone_catalogue WHERE price_category IS NOT NULL ORDER BY price_category")
    pcats = [r["price_category"] for r in (cur.fetchall() or [])]
    cur.execute("SELECT DISTINCT network_type FROM phone_catalogue WHERE network_type IS NOT NULL ORDER BY network_type")
    networks = [r["network_type"] for r in (cur.fetchall() or [])]

    cur.close(); conn.close()

    pages = (total + PAGE_SIZE - 1) // PAGE_SIZE

    return render_template(
        "portal/admin_catalogue.html",
        rows=rows, total=total, page=page, pages=pages,
        brands=brands, pcats=pcats, networks=networks,
        q=q, brand=brand, price_category=pcat,
        network_type=network, nfc=nfc, is_active=active,
    )


@portal_admin_bp.route("/catalogue/phones/<int:phone_id>/edit", methods=["GET", "POST"])
def catalogue_edit(phone_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        data = request.form
        cur.execute(
            """UPDATE phone_catalogue SET
                brand=%s, model_name=%s, variant_name=%s, release_year=%s,
                price_category=%s, network_type=%s, screen_size_inches=%s,
                display_type=%s, refresh_rate_hz=%s, screen_resolution=%s,
                chipset_model=%s, ram=%s, storage=%s, battery_capacity_mah=%s,
                fast_charging_watts=%s, rear_camera_main_mp=%s, front_camera_mp=%s,
                video_recording=%s, gaming_rating=%s, battery_performance=%s,
                camera_quality_rating=%s, wifi_version=%s, bluetooth_version=%s,
                nfc=%s, body_material=%s, water_resistance=%s, fingerprint_type=%s,
                available_colors=%s, nigeria_market_price_naira=%s, best_for=%s,
                search_intent_tags=%s, ai_summary=%s, ai_sales_pitch=%s,
                is_active=%s, updated_at=NOW()
               WHERE id=%s""",
            (
                data.get("brand") or None,
                data.get("model_name") or None,
                data.get("variant_name") or None,
                int(data["release_year"]) if data.get("release_year") else None,
                data.get("price_category") or None,
                data.get("network_type") or None,
                float(data["screen_size_inches"]) if data.get("screen_size_inches") else None,
                data.get("display_type") or None,
                int(data["refresh_rate_hz"]) if data.get("refresh_rate_hz") else None,
                data.get("screen_resolution") or None,
                data.get("chipset_model") or None,
                data.get("ram") or None,
                data.get("storage") or None,
                int(data["battery_capacity_mah"]) if data.get("battery_capacity_mah") else None,
                int(data["fast_charging_watts"]) if data.get("fast_charging_watts") else None,
                int(data["rear_camera_main_mp"]) if data.get("rear_camera_main_mp") else None,
                int(data["front_camera_mp"]) if data.get("front_camera_mp") else None,
                data.get("video_recording") or None,
                data.get("gaming_rating") or None,
                data.get("battery_performance") or None,
                data.get("camera_quality_rating") or None,
                data.get("wifi_version") or None,
                data.get("bluetooth_version") or None,
                data.get("nfc") or None,
                data.get("body_material") or None,
                data.get("water_resistance") or None,
                data.get("fingerprint_type") or None,
                data.get("available_colors") or None,
                float(data["nigeria_market_price_naira"]) if data.get("nigeria_market_price_naira") else None,
                data.get("best_for") or None,
                data.get("search_intent_tags") or None,
                data.get("ai_summary") or None,
                data.get("ai_sales_pitch") or None,
                "is_active" in data,
                phone_id,
            ),
        )
        conn.commit()
        insert_audit_log(action="catalogue_edit", admin_username=_admin_user(),
                         details={"phone_id": phone_id})
        cur.close(); conn.close()
        flash("Phone updated.", "success")
        return redirect(url_for("portal_admin.catalogue"))

    cur.execute("SELECT * FROM phone_catalogue WHERE id=%s", (phone_id,))
    phone = cur.fetchone()
    cur.close(); conn.close()

    if not phone:
        flash("Phone not found.", "danger")
        return redirect(url_for("portal_admin.catalogue"))

    return render_template("portal/admin_catalogue_edit.html", phone=phone)


@portal_admin_bp.route("/catalogue/phones/<int:phone_id>/delete", methods=["POST"])
def catalogue_delete(phone_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM phone_catalogue WHERE id=%s", (phone_id,))
    conn.commit()
    cur.close(); conn.close()

    insert_audit_log(action="catalogue_delete", admin_username=_admin_user(),
                     details={"phone_id": phone_id})
    flash("Phone deleted.", "success")
    return redirect(url_for("portal_admin.catalogue"))


@portal_admin_bp.route("/catalogue/phones/bulk", methods=["POST"])
def catalogue_bulk():
    r = _require_admin()
    if r: return r

    action  = request.form.get("bulk_action", "")
    ids_raw = request.form.getlist("selected_ids")
    ids     = [int(i) for i in ids_raw if i.isdigit()]

    if not ids:
        flash("No rows selected.", "warning")
        return redirect(url_for("portal_admin.catalogue"))

    conn = get_db_connection()
    cur  = conn.cursor()

    if action == "delete":
        cur.execute("DELETE FROM phone_catalogue WHERE id = ANY(%s)", (ids,))
        conn.commit()
        insert_audit_log(action="catalogue_bulk_delete", admin_username=_admin_user(),
                         details={"count": len(ids)})
        flash(f"{len(ids)} phone(s) deleted.", "success")

    elif action == "activate":
        cur.execute("UPDATE phone_catalogue SET is_active=TRUE, updated_at=NOW() WHERE id = ANY(%s)", (ids,))
        conn.commit()
        flash(f"{len(ids)} phone(s) activated.", "success")

    elif action == "deactivate":
        cur.execute("UPDATE phone_catalogue SET is_active=FALSE, updated_at=NOW() WHERE id = ANY(%s)", (ids,))
        conn.commit()
        flash(f"{len(ids)} phone(s) deactivated.", "success")

    elif action == "set_field":
        field  = request.form.get("bulk_field", "").strip()
        value  = request.form.get("bulk_value", "").strip()
        allowed = {"brand","price_category","network_type","display_type","nfc",
                   "body_material","water_resistance","fingerprint_type","gaming_rating",
                   "battery_performance","camera_quality_rating","best_for"}
        if field not in allowed:
            flash("Invalid field for bulk update.", "danger")
        else:
            cur.execute(
                f"UPDATE phone_catalogue SET {field}=%s, updated_at=NOW() WHERE id = ANY(%s)",
                (value or None, ids),
            )
            conn.commit()
            insert_audit_log(action="catalogue_bulk_field_update", admin_username=_admin_user(),
                             details={"field": field, "value": value, "count": len(ids)})
            flash(f"{len(ids)} phone(s) updated — {field} set to '{value}'.", "success")

    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue"))


@portal_admin_bp.route("/catalogue/phones/export")
def catalogue_export():
    r = _require_admin()
    if r: return r

    q       = (request.args.get("q") or "").strip()
    brand   = (request.args.get("brand") or "").strip()
    pcat    = (request.args.get("price_category") or "").strip()
    network = (request.args.get("network_type") or "").strip()
    nfc     = (request.args.get("nfc") or "").strip()
    active  = request.args.get("is_active", "")

    conditions, params = [], []
    if q:
        conditions.append(
            "to_tsvector('english', COALESCE(brand,'') || ' ' || COALESCE(model_name,'') || ' ' || COALESCE(variant_name,'') || ' ' || COALESCE(search_intent_tags,'')) @@ plainto_tsquery('english', %s)"
        )
        params.append(q)
    if brand:
        conditions.append("brand = %s"); params.append(brand)
    if pcat:
        conditions.append("price_category = %s"); params.append(pcat)
    if network:
        conditions.append("network_type = %s"); params.append(network)
    if nfc:
        conditions.append("nfc = %s"); params.append(nfc)
    if active in ("true", "false"):
        conditions.append("is_active = %s"); params.append(active == "true")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT {', '.join(_CAT_COLS)} FROM phone_catalogue {where} ORDER BY brand, model_name, variant_name",
        params,
    )
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=_CAT_COLS)
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in _CAT_COLS})

    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=phone_catalogue_export.csv"},
    )


@portal_admin_bp.route("/catalogue/phones/import", methods=["GET", "POST"])
def catalogue_import():
    r = _require_admin()
    if r: return r

    if request.method == "GET":
        return render_template("portal/admin_catalogue_import.html")

    file = request.files.get("file")
    if not file or not file.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("portal_admin.catalogue_import"))

    filename = file.filename.lower()
    inserted = updated = errors = 0

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        if filename.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            rows_iter = reader
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows_iter = (dict(zip(headers, [c for c in row])) for row in ws.iter_rows(min_row=2, values_only=True))
        else:
            flash("Only .csv or .xlsx files are supported.", "danger")
            cur.close(); conn.close()
            return redirect(url_for("portal_admin.catalogue_import"))

        for row in rows_iter:
            pid = str(row.get("product_id") or "").strip()
            if not pid:
                errors += 1
                continue
            try:
                cur.execute(
                    """INSERT INTO phone_catalogue (
                        product_id,brand,model_name,variant_name,release_year,price_category,
                        network_type,screen_size_inches,display_type,refresh_rate_hz,screen_resolution,
                        chipset_model,ram,storage,battery_capacity_mah,fast_charging_watts,
                        rear_camera_main_mp,front_camera_mp,video_recording,gaming_rating,
                        battery_performance,camera_quality_rating,wifi_version,bluetooth_version,
                        nfc,body_material,water_resistance,fingerprint_type,available_colors,
                        nigeria_market_price_naira,best_for,search_intent_tags,ai_summary,ai_sales_pitch
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (product_id) DO UPDATE SET
                        brand=%s,model_name=%s,variant_name=%s,updated_at=NOW()""",
                    (
                        pid,
                        row.get("brand"), row.get("model_name"), row.get("variant_name"),
                        int(row["release_year"]) if row.get("release_year") else None,
                        row.get("price_category"), row.get("network_type"),
                        float(row["screen_size_inches"]) if row.get("screen_size_inches") else None,
                        row.get("display_type"),
                        int(row["refresh_rate_hz"]) if row.get("refresh_rate_hz") else None,
                        row.get("screen_resolution"), row.get("chipset_model"),
                        row.get("ram"), row.get("storage"),
                        int(row["battery_capacity_mah"]) if row.get("battery_capacity_mah") else None,
                        int(row["fast_charging_watts"]) if row.get("fast_charging_watts") else None,
                        int(row["rear_camera_main_mp"]) if row.get("rear_camera_main_mp") else None,
                        int(row["front_camera_mp"]) if row.get("front_camera_mp") else None,
                        row.get("video_recording"), row.get("gaming_rating"),
                        row.get("battery_performance"), row.get("camera_quality_rating"),
                        row.get("wifi_version"), row.get("bluetooth_version"), row.get("nfc"),
                        row.get("body_material"), row.get("water_resistance"),
                        row.get("fingerprint_type"), row.get("available_colors"),
                        float(row["nigeria_market_price_naira"]) if row.get("nigeria_market_price_naira") else None,
                        row.get("best_for"), row.get("search_intent_tags"),
                        row.get("ai_summary"), row.get("ai_sales_pitch"),
                        # ON CONFLICT update fields
                        row.get("brand"), row.get("model_name"), row.get("variant_name"),
                    ),
                )
                if cur.rowcount == 1:
                    inserted += 1
                else:
                    updated += 1
            except Exception:
                errors += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f"Import failed: {e}", "danger")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_import"))

    cur.close(); conn.close()
    insert_audit_log(action="catalogue_import", admin_username=_admin_user(),
                     details={"inserted": inserted, "updated": updated, "errors": errors})
    flash(f"Import complete — {inserted} added, {updated} updated, {errors} skipped.", "success")
    return redirect(url_for("portal_admin.catalogue"))


# ══════════════════════════════════════════════════════════════════════════════
# PHONE BRANDS
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/brands")
def brands():
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT pb.id, pb.brand, pb.priority, pb.created_at, pb.updated_at,
               COUNT(pc.id) AS phone_count
        FROM phone_brands pb
        LEFT JOIN phone_catalogue pc ON pc.brand = pb.brand
        GROUP BY pb.id, pb.brand, pb.priority, pb.created_at, pb.updated_at
        ORDER BY pb.brand
    """)
    rows = cur.fetchall() or []
    cur.close(); conn.close()
    return render_template("portal/admin_brands.html", brands=rows)


@portal_admin_bp.route("/brands/add", methods=["POST"])
def brands_add():
    r = _require_admin()
    if r: return r

    brand    = (request.form.get("brand") or "").strip()
    priority = (request.form.get("priority") or "Medium").strip()

    if not brand:
        flash("Brand name is required.", "danger")
        return redirect(url_for("portal_admin.brands"))

    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO phone_brands (brand, priority) VALUES (%s, %s) ON CONFLICT (brand) DO NOTHING",
            (brand, priority),
        )
        conn.commit()
        flash(f"Brand '{brand}' added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    finally:
        cur.close(); conn.close()

    return redirect(url_for("portal_admin.brands"))


@portal_admin_bp.route("/brands/<int:brand_id>/edit", methods=["GET", "POST"])
def brands_edit(brand_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        brand    = (request.form.get("brand") or "").strip()
        priority = (request.form.get("priority") or "Medium").strip()
        cur2 = conn.cursor()
        cur2.execute(
            "UPDATE phone_brands SET brand=%s, priority=%s, updated_at=NOW() WHERE id=%s",
            (brand, priority, brand_id),
        )
        conn.commit()
        cur2.close(); cur.close(); conn.close()
        flash("Brand updated.", "success")
        return redirect(url_for("portal_admin.brands"))

    cur.execute("SELECT * FROM phone_brands WHERE id=%s", (brand_id,))
    brand_row = cur.fetchone()
    cur.close(); conn.close()

    if not brand_row:
        flash("Brand not found.", "danger")
        return redirect(url_for("portal_admin.brands"))

    return render_template("portal/admin_brands.html", edit_brand=brand_row, brands=[])


@portal_admin_bp.route("/brands/<int:brand_id>/delete", methods=["POST"])
def brands_delete(brand_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM phone_brands WHERE id=%s", (brand_id,))
    conn.commit()
    cur.close(); conn.close()
    flash("Brand deleted.", "success")
    return redirect(url_for("portal_admin.brands"))


# ══════════════════════════════════════════════════════════════════════════════
# MULTI-CATEGORY CATALOGUE
# ══════════════════════════════════════════════════════════════════════════════

import re as _re
import json as _json_cat

def _slugify(text: str) -> str:
    return _re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')


def _cat_get(cur, category_id: int):
    cur.execute("SELECT * FROM catalogue_categories WHERE id=%s", (category_id,))
    return cur.fetchone()


def _cat_attrs(cur, category_id: int):
    cur.execute(
        "SELECT * FROM catalogue_attribute_definitions "
        "WHERE category_id=%s ORDER BY sort_order, id",
        (category_id,)
    )
    return cur.fetchall()


def _cat_variant_types(cur, category_id: int):
    cur.execute("""
        SELECT vt.*, json_agg(
            json_build_object('id', vo.id, 'value', vo.value, 'sort_order', vo.sort_order)
            ORDER BY vo.sort_order, vo.id
        ) FILTER (WHERE vo.id IS NOT NULL) AS options
        FROM catalogue_variant_types vt
        LEFT JOIN catalogue_variant_options vo ON vo.variant_type_id = vt.id
        WHERE vt.category_id = %s
        GROUP BY vt.id
        ORDER BY vt.sort_order, vt.id
    """, (category_id,))
    return cur.fetchall()


def _product_variants(cur, product_id: int):
    cur.execute("""
        SELECT * FROM catalogue_product_variants
        WHERE product_id = %s
        ORDER BY variant_combo::text
    """, (product_id,))
    return cur.fetchall()


# ── Department helpers ────────────────────────────────────────────────────────

def _dept_get(cur, dept_id: int):
    cur.execute("SELECT * FROM catalogue_departments WHERE id=%s", (dept_id,))
    return cur.fetchone()

def _all_departments(cur):
    cur.execute("SELECT * FROM catalogue_departments ORDER BY sort_order, name")
    return cur.fetchall()

DEPT_ICONS = ["🏪","📱","💊","💄","🛒","🖨","🛋","👗","🔧","🍔","🚗","🏋","📚","🌿","🐾","🧸","⚗️","🏠","🎓","💈"]


# ── Department CRUD ───────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/departments/new", methods=["GET", "POST"])
def catalogue_department_new():
    r = _require_admin()
    if r: return r

    if request.method == "GET":
        return render_template("portal/admin_department_new.html", icons=DEPT_ICONS)

    name        = (request.form.get("name") or "").strip()
    icon        = (request.form.get("icon") or "🏪").strip()
    description = (request.form.get("description") or "").strip()

    if not name:
        flash("Department name is required.", "danger")
        return render_template("portal/admin_department_new.html", icons=DEPT_ICONS)

    slug = _slugify(name)
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO catalogue_departments (name, slug, icon, description, created_by)
               VALUES (%s, %s, %s, %s, %s) RETURNING id""",
            (name, slug, icon, description or None, _admin_user())
        )
        conn.commit()
        insert_audit_log(action="catalogue_department_create", admin_username=_admin_user(),
                         details={"name": name, "slug": slug})
        flash(f"Department '{name}' created.", "success")
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        flash(f"A department with the slug '{slug}' already exists.", "danger")
        cur.close(); conn.close()
        return render_template("portal/admin_department_new.html", icons=DEPT_ICONS)
    finally:
        cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_dashboard"))


@portal_admin_bp.route("/catalogue/departments/<int:dept_id>/edit", methods=["GET", "POST"])
def catalogue_department_edit(dept_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    dept = _dept_get(cur, dept_id)
    if not dept:
        cur.close(); conn.close()
        flash("Department not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    if request.method == "POST":
        name        = (request.form.get("name") or "").strip()
        icon        = (request.form.get("icon") or dept["icon"]).strip()
        description = (request.form.get("description") or "").strip()
        if not name:
            flash("Name is required.", "danger")
        else:
            cur.execute(
                "UPDATE catalogue_departments SET name=%s, icon=%s, description=%s WHERE id=%s",
                (name, icon, description or None, dept_id)
            )
            conn.commit()
            insert_audit_log(action="catalogue_department_edit", admin_username=_admin_user(),
                             details={"id": dept_id, "name": name})
            flash("Department updated.", "success")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    cur.close(); conn.close()
    return render_template("portal/admin_department_new.html", icons=DEPT_ICONS, dept=dept)


@portal_admin_bp.route("/catalogue/departments/<int:dept_id>/toggle", methods=["POST"])
def catalogue_department_toggle(dept_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("UPDATE catalogue_departments SET is_active = NOT is_active WHERE id=%s", (dept_id,))
    conn.commit()
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_dashboard"))


@portal_admin_bp.route("/catalogue/departments/<int:dept_id>/delete", methods=["POST"])
def catalogue_department_delete(dept_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Only allow delete if no categories assigned
    cur.execute("SELECT COUNT(*) AS n FROM catalogue_categories WHERE department_id=%s", (dept_id,))
    count = (cur.fetchone() or {}).get("n", 0)
    if count > 0:
        flash(f"Cannot delete: {count} category/categories are still assigned to this department.", "danger")
    else:
        cur.execute("DELETE FROM catalogue_departments WHERE id=%s", (dept_id,))
        conn.commit()
        insert_audit_log(action="catalogue_department_delete", admin_username=_admin_user(),
                         details={"id": dept_id})
        flash("Department deleted.", "success")
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_dashboard"))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue")
def catalogue_dashboard():
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # All departments ordered
    cur.execute("SELECT * FROM catalogue_departments ORDER BY sort_order, name")
    departments = cur.fetchall()

    # All categories with product counts and department info
    cur.execute("""
        SELECT c.*, d.name AS dept_name, COUNT(p.id) AS product_count
        FROM catalogue_categories c
        LEFT JOIN catalogue_departments d ON d.id = c.department_id
        LEFT JOIN catalogue_products p ON p.category_id = c.id
        GROUP BY c.id, d.name
        ORDER BY c.sort_order, c.name
    """)
    all_categories = cur.fetchall()

    # Group categories by department_id for template rendering
    import collections
    cats_by_dept = collections.defaultdict(list)
    ungrouped    = []
    for cat in all_categories:
        if cat["department_id"]:
            cats_by_dept[cat["department_id"]].append(cat)
        else:
            ungrouped.append(cat)

    # Legacy phones count
    cur.execute("SELECT COUNT(*) AS n FROM phone_catalogue")
    phones_count = (cur.fetchone() or {}).get("n", 0)

    # Recent uploads (last 10)
    cur.execute("""
        SELECT u.*, c.name AS category_name
        FROM catalogue_uploads u
        LEFT JOIN catalogue_categories c ON c.id = u.category_id
        ORDER BY u.created_at DESC LIMIT 10
    """)
    recent_uploads = cur.fetchall()

    cur.close(); conn.close()
    return render_template(
        "portal/admin_catalogue_dashboard.html",
        departments=departments,
        cats_by_dept=dict(cats_by_dept),
        ungrouped=ungrouped,
        phones_count=phones_count,
        recent_uploads=recent_uploads,
    )


# ── Create category ───────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/new", methods=["GET", "POST"])
def catalogue_category_new():
    r = _require_admin()
    if r: return r

    ICONS = ["📱","💻","🖥","🖵","🔌","📷","🎮","📺","🎧","⌨️","🖱","🔋","📡","🖨","⌚","📻",
             "💊","💄","🛒","🛋","👗","🔧","🍔","🏋","📚","🌿"]

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    departments = _all_departments(cur)
    cur.close(); conn.close()

    if request.method == "GET":
        return render_template("portal/admin_category_new.html", icons=ICONS, departments=departments)

    name          = (request.form.get("name") or "").strip()
    icon          = (request.form.get("icon") or "📦").strip()
    description   = (request.form.get("description") or "").strip()
    dept_id_raw   = request.form.get("department_id") or ""
    department_id = int(dept_id_raw) if dept_id_raw.isdigit() else None

    if not name:
        flash("Category name is required.", "danger")
        return render_template("portal/admin_category_new.html", icons=ICONS, departments=departments)

    slug = _slugify(name)
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO catalogue_categories (name, slug, icon, description, department_id, created_by)
               VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
            (name, slug, icon, description or None, department_id, _admin_user())
        )
        new_id = cur.fetchone()["id"]
        conn.commit()
        insert_audit_log(action="catalogue_category_create", admin_username=_admin_user(),
                         details={"name": name, "slug": slug, "department_id": department_id})
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        flash(f"A category with the slug '{slug}' already exists.", "danger")
        cur.close(); conn.close()
        return render_template("portal/admin_category_new.html", icons=ICONS, departments=departments)
    finally:
        cur.close(); conn.close()

    flash(f"Category '{name}' created. Now define its attributes.", "success")
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=new_id))


# ── Edit category ─────────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/edit", methods=["GET", "POST"])
def catalogue_category_edit(category_id: int):
    r = _require_admin()
    if r: return r

    ICONS = ["📱","💻","🖥","🖵","🔌","📷","🎮","📺","🎧","⌨️","🖱","🔋","📡","🖨","⌚","📻",
             "💊","💄","🛒","🛋","👗","🔧","🍔","🏋","📚","🌿"]
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    departments = _all_departments(cur)

    if request.method == "POST":
        name          = (request.form.get("name") or "").strip()
        icon          = (request.form.get("icon") or cat["icon"]).strip()
        description   = (request.form.get("description") or "").strip()
        dept_id_raw   = request.form.get("department_id") or ""
        department_id = int(dept_id_raw) if dept_id_raw.isdigit() else None
        if not name:
            flash("Name is required.", "danger")
        else:
            try:
                cur.execute(
                    "UPDATE catalogue_categories SET name=%s, icon=%s, description=%s, department_id=%s WHERE id=%s",
                    (name, icon, description or None, department_id, category_id)
                )
                conn.commit()
                insert_audit_log(action="catalogue_category_edit", admin_username=_admin_user(),
                                 details={"id": category_id, "name": name, "department_id": department_id})
                flash("Category updated.", "success")
            except Exception as e:
                conn.rollback()
                flash(f"Error: {e}", "danger")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))

    cur.close(); conn.close()
    return render_template("portal/admin_category_new.html", icons=ICONS, cat=cat, departments=departments)


# ── Toggle category active ────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/toggle", methods=["POST"])
def catalogue_category_toggle(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute(
        "UPDATE catalogue_categories SET is_active = NOT is_active WHERE id=%s",
        (category_id,)
    )
    conn.commit()
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_dashboard"))


# ── Delete category ───────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/delete", methods=["POST"])
def catalogue_category_delete(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT COUNT(*) AS n FROM catalogue_products WHERE category_id=%s", (category_id,))
    count = (cur.fetchone() or {}).get("n", 0)
    if count > 0:
        cur.close(); conn.close()
        flash(f"Cannot delete: category has {count} products. Deactivate it instead.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    cur.execute("DELETE FROM catalogue_categories WHERE id=%s", (category_id,))
    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="catalogue_category_delete", admin_username=_admin_user(),
                     details={"id": category_id})
    flash("Category deleted.", "success")
    return redirect(url_for("portal_admin.catalogue_dashboard"))


# ── Manage attributes ─────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/attributes", methods=["GET", "POST"])
def catalogue_category_attributes(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "add":
            label = (request.form.get("attribute_label") or "").strip()
            key   = _slugify(label).replace("-", "_") if label else ""
            dtype = request.form.get("data_type", "text")
            unit  = (request.form.get("unit") or "").strip() or None
            filterable = request.form.get("is_filterable") == "1"
            required   = request.form.get("is_required") == "1"
            if not label:
                flash("Attribute label is required.", "danger")
            else:
                # Get next sort order
                cur.execute(
                    "SELECT COALESCE(MAX(sort_order),0)+1 AS n FROM catalogue_attribute_definitions WHERE category_id=%s",
                    (category_id,)
                )
                sort_order = (cur.fetchone() or {}).get("n", 1)
                try:
                    cur.execute(
                        """INSERT INTO catalogue_attribute_definitions
                           (category_id, attribute_key, attribute_label, data_type, unit,
                            is_filterable, is_required, sort_order)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (category_id, key, label, dtype, unit, filterable, required, sort_order)
                    )
                    conn.commit()
                    flash(f"Attribute '{label}' added.", "success")
                except psycopg2.errors.UniqueViolation:
                    conn.rollback()
                    flash(f"Attribute key '{key}' already exists in this category.", "danger")

        elif action == "delete":
            attr_id = request.form.get("attr_id")
            if attr_id:
                cur.execute("DELETE FROM catalogue_attribute_definitions WHERE id=%s AND category_id=%s",
                            (int(attr_id), category_id))
                conn.commit()
                flash("Attribute removed.", "success")

        elif action == "toggle_required":
            attr_id = request.form.get("attr_id")
            if attr_id:
                cur.execute(
                    "UPDATE catalogue_attribute_definitions SET is_required = NOT is_required WHERE id=%s AND category_id=%s",
                    (int(attr_id), category_id)
                )
                conn.commit()

        elif action == "toggle_filterable":
            attr_id = request.form.get("attr_id")
            if attr_id:
                cur.execute(
                    "UPDATE catalogue_attribute_definitions SET is_filterable = NOT is_filterable WHERE id=%s AND category_id=%s",
                    (int(attr_id), category_id)
                )
                conn.commit()

        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))

    attrs         = _cat_attrs(cur, category_id)
    variant_types = _cat_variant_types(cur, category_id)

    # Load templates — prefer those matching the category's department
    conn2 = get_db_connection()
    cur2  = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur2.execute("""
        SELECT t.id, t.name, t.slug, t.is_builtin,
               jsonb_array_length(t.attributes) AS attr_count
        FROM catalogue_industry_templates t
        ORDER BY
            CASE WHEN t.department_id = %s THEN 0 ELSE 1 END,
            t.is_builtin DESC, t.name
    """, (cat.get("department_id"),))
    templates = cur2.fetchall()
    cur2.close(); conn2.close()

    cur.close(); conn.close()
    return render_template(
        "portal/admin_category_attributes.html",
        cat=cat, attrs=attrs,
        DATA_TYPES=["text", "number", "select", "boolean"],
        templates=templates,
        variant_types=variant_types,
    )


# ── Apply template to category ────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/apply-template", methods=["POST"])
def catalogue_apply_template(category_id: int):
    r = _require_admin()
    if r: return r

    tpl_id_raw = request.form.get("template_id") or ""
    if not tpl_id_raw.isdigit():
        flash("Please select a template.", "warning")
        return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cat = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    cur.execute("SELECT * FROM catalogue_industry_templates WHERE id=%s", (int(tpl_id_raw),))
    tpl = cur.fetchone()
    if not tpl:
        cur.close(); conn.close()
        flash("Template not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))

    # Get current max sort_order
    cur.execute(
        "SELECT COALESCE(MAX(sort_order),0) AS n FROM catalogue_attribute_definitions WHERE category_id=%s",
        (category_id,)
    )
    base_sort = (cur.fetchone() or {}).get("n", 0)

    # Get existing attribute keys to avoid duplicates
    cur.execute(
        "SELECT attribute_key FROM catalogue_attribute_definitions WHERE category_id=%s",
        (category_id,)
    )
    existing_keys = {r["attribute_key"] for r in cur.fetchall()}

    import json as _json
    attrs = tpl["attributes"] if isinstance(tpl["attributes"], list) else _json.loads(tpl["attributes"])
    added = 0
    for a in attrs:
        key = a.get("key", "")
        if not key or key in existing_keys:
            continue
        cur.execute("""
            INSERT INTO catalogue_attribute_definitions
                (category_id, attribute_key, attribute_label, data_type, unit,
                 is_filterable, is_required, sort_order)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (category_id, attribute_key) DO NOTHING
        """, (
            category_id, key, a.get("label", key),
            a.get("data_type", "text"), a.get("unit") or None,
            bool(a.get("is_filterable", False)), bool(a.get("is_required", False)),
            base_sort + a.get("sort_order", 99)
        ))
        added += 1

    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="catalogue_apply_template", admin_username=_admin_user(),
                     details={"category_id": category_id, "template_id": int(tpl_id_raw),
                              "template_name": tpl["name"], "added": added})
    flash(f"Template '{tpl['name']}' applied — {added} attribute(s) added.", "success")
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))


# ── Industry templates management ─────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/templates")
def catalogue_templates():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT t.*, d.name AS dept_name, d.icon AS dept_icon,
               jsonb_array_length(t.attributes) AS attr_count
        FROM catalogue_industry_templates t
        LEFT JOIN catalogue_departments d ON d.id = t.department_id
        ORDER BY t.is_builtin DESC, t.name
    """)
    templates = cur.fetchall()
    cur.close(); conn.close()
    return render_template("portal/admin_templates.html", templates=templates)


@portal_admin_bp.route("/catalogue/templates/new", methods=["GET", "POST"])
def catalogue_template_new():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    departments = _all_departments(cur)

    if request.method == "POST":
        name       = (request.form.get("name") or "").strip()
        dept_raw   = request.form.get("department_id") or ""
        dept_id    = int(dept_raw) if dept_raw.isdigit() else None
        # Collect attribute rows from form
        import json as _json
        keys    = request.form.getlist("attr_key")
        labels  = request.form.getlist("attr_label")
        dtypes  = request.form.getlist("attr_data_type")
        units   = request.form.getlist("attr_unit")
        filts   = set(request.form.getlist("attr_filterable"))
        reqs    = set(request.form.getlist("attr_required"))
        attrs   = []
        for i, (k, l) in enumerate(zip(keys, labels)):
            k = k.strip(); l = l.strip()
            if not k or not l:
                continue
            attrs.append({
                "key": k, "label": l,
                "data_type": dtypes[i] if i < len(dtypes) else "text",
                "unit": units[i].strip() if i < len(units) else "",
                "is_filterable": str(i) in filts,
                "is_required": str(i) in reqs,
                "sort_order": i + 1,
            })
        if not name:
            flash("Template name is required.", "danger")
        elif not attrs:
            flash("Add at least one attribute.", "danger")
        else:
            slug = _slugify(name)
            try:
                cur.execute("""
                    INSERT INTO catalogue_industry_templates
                        (name, slug, department_id, attributes, is_builtin, created_by)
                    VALUES (%s,%s,%s,%s,FALSE,%s) RETURNING id
                """, (name, slug, dept_id, _json.dumps(attrs), _admin_user()))
                conn.commit()
                insert_audit_log(action="catalogue_template_create", admin_username=_admin_user(),
                                 details={"name": name})
                flash(f"Template '{name}' created.", "success")
                cur.close(); conn.close()
                return redirect(url_for("portal_admin.catalogue_templates"))
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                flash(f"A template named '{name}' already exists.", "danger")

    cur.close(); conn.close()
    return render_template("portal/admin_template_edit.html",
                           departments=departments, tpl=None,
                           DATA_TYPES=["text", "number", "select", "boolean"])


@portal_admin_bp.route("/catalogue/templates/<int:tpl_id>/edit", methods=["GET", "POST"])
def catalogue_template_edit(tpl_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM catalogue_industry_templates WHERE id=%s", (tpl_id,))
    tpl = cur.fetchone()
    if not tpl:
        cur.close(); conn.close()
        flash("Template not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_templates"))
    departments = _all_departments(cur)

    if request.method == "POST":
        if tpl["is_builtin"]:
            cur.close(); conn.close()
            flash("Built-in templates cannot be edited.", "danger")
            return redirect(url_for("portal_admin.catalogue_templates"))
        import json as _json
        name     = (request.form.get("name") or "").strip()
        dept_raw = request.form.get("department_id") or ""
        dept_id  = int(dept_raw) if dept_raw.isdigit() else None
        keys    = request.form.getlist("attr_key")
        labels  = request.form.getlist("attr_label")
        dtypes  = request.form.getlist("attr_data_type")
        units   = request.form.getlist("attr_unit")
        filts   = set(request.form.getlist("attr_filterable"))
        reqs    = set(request.form.getlist("attr_required"))
        attrs   = []
        for i, (k, l) in enumerate(zip(keys, labels)):
            k = k.strip(); l = l.strip()
            if not k or not l:
                continue
            attrs.append({
                "key": k, "label": l,
                "data_type": dtypes[i] if i < len(dtypes) else "text",
                "unit": units[i].strip() if i < len(units) else "",
                "is_filterable": str(i) in filts,
                "is_required": str(i) in reqs,
                "sort_order": i + 1,
            })
        if not name:
            flash("Template name is required.", "danger")
        elif not attrs:
            flash("Add at least one attribute.", "danger")
        else:
            cur.execute("""
                UPDATE catalogue_industry_templates
                SET name=%s, department_id=%s, attributes=%s
                WHERE id=%s
            """, (name, dept_id, _json.dumps(attrs), tpl_id))
            conn.commit()
            insert_audit_log(action="catalogue_template_edit", admin_username=_admin_user(),
                             details={"id": tpl_id, "name": name})
            flash(f"Template '{name}' updated.", "success")
            cur.close(); conn.close()
            return redirect(url_for("portal_admin.catalogue_templates"))

    cur.close(); conn.close()
    return render_template("portal/admin_template_edit.html",
                           departments=departments, tpl=tpl,
                           DATA_TYPES=["text", "number", "select", "boolean"])


@portal_admin_bp.route("/catalogue/templates/<int:tpl_id>/delete", methods=["POST"])
def catalogue_template_delete(tpl_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT name, is_builtin FROM catalogue_industry_templates WHERE id=%s", (tpl_id,))
    tpl = cur.fetchone()
    if tpl and tpl["is_builtin"]:
        flash("Built-in templates cannot be deleted.", "danger")
    elif tpl:
        cur.execute("DELETE FROM catalogue_industry_templates WHERE id=%s", (tpl_id,))
        conn.commit()
        flash(f"Template '{tpl['name']}' deleted.", "success")
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_templates"))


# ══════════════════════════════════════════════════════════════════════════════
# VARIANT TYPES & OPTIONS (per category)
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/variants/add", methods=["POST"])
def catalogue_variant_type_add(category_id: int):
    r = _require_admin()
    if r: return r
    name = (request.form.get("variant_name") or "").strip()
    if not name:
        flash("Variant type name is required.", "danger")
        return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO catalogue_variant_types (category_id, name, sort_order)
            VALUES (%s, %s, (SELECT COALESCE(MAX(sort_order),0)+1 FROM catalogue_variant_types WHERE category_id=%s))
        """, (category_id, name, category_id))
        conn.commit()
        flash(f"Variant type '{name}' added.", "success")
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        flash(f"Variant type '{name}' already exists in this category.", "danger")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))


@portal_admin_bp.route("/catalogue/categories/<int:category_id>/variants/<int:type_id>/delete", methods=["POST"])
def catalogue_variant_type_delete(category_id: int, type_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM catalogue_variant_types WHERE id=%s AND category_id=%s", (type_id, category_id))
    # Purge saved variant rows for all products in this category — JSONB combos are now stale
    cur.execute("""
        DELETE FROM catalogue_product_variants
        WHERE product_id IN (SELECT id FROM catalogue_products WHERE category_id=%s)
    """, (category_id,))
    conn.commit()
    cur.close(); conn.close()
    flash("Variant type and all its options deleted. Saved variant rows cleared — please re-save product variants.", "success")
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))


@portal_admin_bp.route("/catalogue/categories/<int:category_id>/variants/<int:type_id>/options/add", methods=["POST"])
def catalogue_variant_option_add(category_id: int, type_id: int):
    r = _require_admin()
    if r: return r
    value = (request.form.get("option_value") or "").strip()
    if not value:
        flash("Option value is required.", "danger")
        return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO catalogue_variant_options (variant_type_id, value, sort_order)
            VALUES (%s, %s, (SELECT COALESCE(MAX(sort_order),0)+1 FROM catalogue_variant_options WHERE variant_type_id=%s))
        """, (type_id, value, type_id))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        flash(f"Option '{value}' already exists.", "danger")
    finally:
        cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))


@portal_admin_bp.route("/catalogue/categories/<int:category_id>/variants/<int:type_id>/options/<int:opt_id>/delete",
                        methods=["POST"])
def catalogue_variant_option_delete(category_id: int, type_id: int, opt_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM catalogue_variant_options WHERE id=%s AND variant_type_id=%s", (opt_id, type_id))
    conn.commit()
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.catalogue_category_attributes", category_id=category_id))


# ── Product variant matrix save ───────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/products/<int:product_id>/variants",
                        methods=["POST"])
def catalogue_product_variants_save(category_id: int, product_id: int):
    r = _require_admin()
    if r: return r

    import json as _json, itertools as _it

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Verify product belongs to category
    cur.execute("SELECT id FROM catalogue_products WHERE id=%s AND category_id=%s", (product_id, category_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        flash("Product not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))

    # Collect variant rows from form: indices 0..N
    indices = sorted(set(
        k.split("_", 2)[2] for k in request.form.keys()
        if k.startswith("v_combo_")
    ), key=lambda x: int(x) if x.isdigit() else 0)

    # Replace the full variant set for this product
    cur.execute("DELETE FROM catalogue_product_variants WHERE product_id=%s", (product_id,))

    saved = 0
    for idx in indices:
        combo_raw = request.form.get(f"v_combo_{idx}", "{}")
        sku_val   = (request.form.get(f"v_sku_{idx}") or "").strip() or None
        price_raw = (request.form.get(f"v_price_{idx}") or "0").strip()
        is_active = request.form.get(f"v_active_{idx}") == "1"
        stock     = request.form.get(f"v_stock_{idx}") or "in_stock"
        try:
            combo = _json.loads(combo_raw)
            price = float(price_raw) if price_raw else 0.0
        except (ValueError, TypeError):
            continue

        cur.execute("SAVEPOINT sp_var")
        try:
            cur.execute("""
                INSERT INTO catalogue_product_variants
                    (product_id, sku, price_modifier, stock_status, is_active, variant_combo)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (product_id, sku_val, price, stock, is_active, _json.dumps(combo)))
            cur.execute("RELEASE SAVEPOINT sp_var")
            saved += 1
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT sp_var")

    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="catalogue_variants_save", admin_username=_admin_user(),
                     details={"product_id": product_id, "saved": saved})
    flash(f"{saved} variant row(s) saved.", "success")
    return redirect(url_for("portal_admin.catalogue_product_edit",
                            category_id=category_id, product_id=product_id))


# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTS FOR A CATEGORY
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/products")
def catalogue_category_products(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    attrs = _cat_attrs(cur, category_id)

    # Filters
    q         = (request.args.get("q") or "").strip()
    brand_f   = (request.args.get("brand") or "").strip()
    status_f  = request.args.get("is_active", "")
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    per_page  = 50

    where  = ["p.category_id = %s"]
    params = [category_id]

    if q:
        where.append("(p.brand ILIKE %s OR p.model_name ILIKE %s OR p.model_number ILIKE %s)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if brand_f:
        where.append("p.brand = %s")
        params.append(brand_f)
    if status_f == "true":
        where.append("p.is_active = TRUE")
    elif status_f == "false":
        where.append("p.is_active = FALSE")

    where_sql = "WHERE " + " AND ".join(where)

    cur.execute(f"SELECT COUNT(*) AS n FROM catalogue_products p {where_sql}", params)
    total = (cur.fetchone() or {}).get("n", 0)
    pages = max(1, (total + per_page - 1) // per_page)
    offset = (page - 1) * per_page

    cur.execute(
        f"SELECT p.* FROM catalogue_products p {where_sql} "
        f"ORDER BY p.brand, p.model_name LIMIT %s OFFSET %s",
        params + [per_page, offset]
    )
    products = cur.fetchall()

    # For each product load its attribute values
    if products:
        pids = [p["id"] for p in products]
        cur.execute(
            """SELECT pa.product_id, ad.attribute_key, pa.value
               FROM catalogue_product_attributes pa
               JOIN catalogue_attribute_definitions ad ON ad.id = pa.attribute_def_id
               WHERE pa.product_id = ANY(%s)""",
            (pids,)
        )
        attr_map: dict = {}
        for row in cur.fetchall():
            attr_map.setdefault(row["product_id"], {})[row["attribute_key"]] = row["value"]
        products = [dict(p, attrs=attr_map.get(p["id"], {})) for p in products]

    # Brand list for filter dropdown
    cur.execute(
        "SELECT DISTINCT brand FROM catalogue_products WHERE category_id=%s AND brand IS NOT NULL ORDER BY brand",
        (category_id,)
    )
    brands = [r["brand"] for r in cur.fetchall()]

    # Department slug for extended column display
    dept_slug = None
    if cat.get("department_id"):
        cur.execute("SELECT slug FROM catalogue_departments WHERE id=%s", (cat["department_id"],))
        dept_row = cur.fetchone()
        dept_slug = (dept_row or {}).get("slug")

    cur.close(); conn.close()
    return render_template(
        "portal/admin_category_products.html",
        cat=cat, attrs=attrs, products=products, brands=brands,
        total=total, page=page, pages=pages, per_page=per_page,
        q=q, brand_f=brand_f, status_f=status_f,
        dept_slug=dept_slug,
    )


# ── Edit a single product ─────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/products/<int:product_id>/edit",
                        methods=["GET", "POST"])
def catalogue_product_edit(category_id: int, product_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    cur.execute("SELECT * FROM catalogue_products WHERE id=%s AND category_id=%s",
                (product_id, category_id))
    product = cur.fetchone()
    if not product:
        cur.close(); conn.close()
        flash("Product not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))

    attrs = _cat_attrs(cur, category_id)

    # Load department slug for field visibility
    dept_slug = None
    if cat.get("department_id"):
        cur.execute("SELECT slug FROM catalogue_departments WHERE id=%s", (cat["department_id"],))
        dept_row = cur.fetchone()
        dept_slug = (dept_row or {}).get("slug")

    # Load existing attribute values
    cur.execute(
        """SELECT ad.attribute_key, pa.value
           FROM catalogue_attribute_definitions ad
           LEFT JOIN catalogue_product_attributes pa
             ON pa.attribute_def_id = ad.id AND pa.product_id = %s
           WHERE ad.category_id = %s ORDER BY ad.sort_order""",
        (product_id, category_id)
    )
    attr_vals = {r["attribute_key"]: r["value"] for r in cur.fetchall()}

    EXTENDED_FIELDS = ["barcode", "unit_of_measure", "weight_value", "weight_unit",
                       "shelf_life_days", "requires_rxn", "regulatory_ref", "dimensions_cm"]

    if request.method == "POST":
        brand        = (request.form.get("brand") or "").strip() or None
        model_name   = (request.form.get("model_name") or "").strip()
        model_number = (request.form.get("model_number") or "").strip() or None
        sku          = (request.form.get("sku") or "").strip() or None
        description  = (request.form.get("description") or "").strip() or None
        image_url    = (request.form.get("image_url") or "").strip() or None
        is_active    = request.form.get("is_active") == "1"

        # Extended fields
        barcode        = (request.form.get("barcode") or "").strip() or None
        unit_of_measure= (request.form.get("unit_of_measure") or "").strip() or None
        weight_value_s = (request.form.get("weight_value") or "").strip()
        try:
            weight_value = float(weight_value_s) if weight_value_s else None
        except ValueError:
            weight_value = None
        weight_unit    = (request.form.get("weight_unit") or "").strip() or None
        shelf_life_s   = (request.form.get("shelf_life_days") or "").strip()
        shelf_life_days= int(shelf_life_s) if shelf_life_s.isdigit() else None
        requires_rxn_s = request.form.get("requires_rxn") or ""
        requires_rxn   = True if requires_rxn_s == "1" else (False if requires_rxn_s == "0" else None)
        regulatory_ref = (request.form.get("regulatory_ref") or "").strip() or None
        dimensions_cm  = (request.form.get("dimensions_cm") or "").strip() or None

        if not model_name:
            flash("Model name is required.", "danger")
            cur.close(); conn.close()
            return render_template(
                "portal/admin_category_product_edit.html",
                cat=cat, product=product, attrs=attrs, dept_slug=dept_slug,
                attr_vals={a["attribute_key"]: request.form.get(f"attr_{a['attribute_key']}") for a in attrs},
            )

        try:
            cur.execute(
                """UPDATE catalogue_products
                   SET brand=%s, model_name=%s, model_number=%s, sku=%s,
                       description=%s, image_url=%s, is_active=%s,
                       barcode=%s, unit_of_measure=%s, weight_value=%s, weight_unit=%s,
                       shelf_life_days=%s, requires_rxn=%s, regulatory_ref=%s,
                       dimensions_cm=%s, updated_at=NOW()
                   WHERE id=%s""",
                (brand, model_name, model_number, sku, description, image_url, is_active,
                 barcode, unit_of_measure, weight_value, weight_unit,
                 shelf_life_days, requires_rxn, regulatory_ref, dimensions_cm,
                 product_id)
            )
            for ad in attrs:
                val = (request.form.get(f"attr_{ad['attribute_key']}") or "").strip() or None
                if val is not None:
                    cur.execute(
                        """INSERT INTO catalogue_product_attributes (product_id, attribute_def_id, value)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (product_id, attribute_def_id) DO UPDATE SET value=EXCLUDED.value""",
                        (product_id, ad["id"], val)
                    )
                else:
                    cur.execute(
                        "DELETE FROM catalogue_product_attributes WHERE product_id=%s AND attribute_def_id=%s",
                        (product_id, ad["id"])
                    )
            conn.commit()
            insert_audit_log(action="catalogue_product_edit", admin_username=_admin_user(),
                             details={"product_id": product_id, "model_name": model_name})
            flash("Product updated.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", "danger")

        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))

    variant_types    = _cat_variant_types(cur, category_id)
    product_variants = _product_variants(cur, product_id)
    cur.close(); conn.close()

    # Build cartesian matrix from variant types × options
    import itertools as _it, json as _json
    existing_by_combo = {
        _json.dumps(dict(sorted(v["variant_combo"].items())), sort_keys=True): v
        for v in product_variants
        if v["variant_combo"]
    }

    matrix = []
    if variant_types:
        type_options = [(vt["name"], vt["options"] or []) for vt in variant_types]
        all_names    = [t[0] for t in type_options]
        all_opts     = [[(o["value"]) for o in t[1]] for t in type_options]
        if all(all_opts):
            for combo_vals in _it.product(*all_opts):
                combo = dict(zip(all_names, combo_vals))
                key   = _json.dumps(combo, sort_keys=True)
                existing = existing_by_combo.get(key)
                matrix.append({
                    "combo":      combo,
                    "combo_json": key,
                    "label":      " / ".join(combo_vals),
                    "sku":        existing["sku"] if existing else "",
                    "price":      float(existing["price_modifier"]) if existing else 0.0,
                    "stock":      existing["stock_status"] if existing else "in_stock",
                    "is_active":  existing["is_active"] if existing else True,
                })

    return render_template(
        "portal/admin_category_product_edit.html",
        cat=cat, product=product, attrs=attrs, attr_vals=attr_vals,
        dept_slug=dept_slug,
        variant_types=variant_types,
        matrix=matrix,
    )


# ── Delete a product ──────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/products/<int:product_id>/delete",
                        methods=["POST"])
def catalogue_product_delete(category_id: int, product_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("DELETE FROM catalogue_products WHERE id=%s AND category_id=%s",
                (product_id, category_id))
    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="catalogue_product_delete", admin_username=_admin_user(),
                     details={"product_id": product_id, "category_id": category_id})
    flash("Product deleted.", "success")
    return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))


# ── Bulk actions ──────────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/products/bulk", methods=["POST"])
def catalogue_products_bulk(category_id: int):
    r = _require_admin()
    if r: return r

    ids_raw = request.form.getlist("ids")
    action  = request.form.get("bulk_action", "")
    ids     = [int(x) for x in ids_raw if x.isdigit()]

    if not ids:
        flash("No products selected.", "warning")
        return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))

    conn = get_db_connection()
    cur  = conn.cursor()
    if action == "activate":
        cur.execute("UPDATE catalogue_products SET is_active=TRUE WHERE id=ANY(%s) AND category_id=%s",
                    (ids, category_id))
    elif action == "deactivate":
        cur.execute("UPDATE catalogue_products SET is_active=FALSE WHERE id=ANY(%s) AND category_id=%s",
                    (ids, category_id))
    elif action == "delete":
        cur.execute("DELETE FROM catalogue_products WHERE id=ANY(%s) AND category_id=%s",
                    (ids, category_id))
    conn.commit()
    cur.close(); conn.close()
    flash(f"Bulk {action} applied to {len(ids)} products.", "success")
    return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))


# ── Download CSV template ─────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/template.csv")
def catalogue_category_template(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    attrs = _cat_attrs(cur, category_id)

    # Determine department slug for extended field inclusion
    dept_slug = None
    if cat.get("department_id"):
        cur.execute("SELECT slug FROM catalogue_departments WHERE id=%s", (cat["department_id"],))
        dept_row = cur.fetchone()
        dept_slug = (dept_row or {}).get("slug")

    cur.close(); conn.close()

    base_cols = ["brand", "model_name", "model_number", "sku", "description", "image_url"]

    # Extended cols selected by department
    _ext_map = {
        "pharmacy":    ["barcode", "unit_of_measure", "shelf_life_days", "regulatory_ref", "requires_rxn"],
        "supermarket": ["barcode", "unit_of_measure", "weight_value", "weight_unit", "shelf_life_days", "regulatory_ref"],
        "beauty":      ["barcode", "unit_of_measure", "weight_value", "weight_unit", "shelf_life_days", "regulatory_ref"],
        "fashion":     ["weight_value", "weight_unit"],
        "furniture":   ["dimensions_cm", "weight_value", "weight_unit"],
        "office":      ["barcode", "dimensions_cm", "weight_value", "weight_unit"],
        "electronics": ["barcode", "dimensions_cm", "weight_value", "weight_unit"],
    }
    ext_cols  = _ext_map.get(dept_slug, []) if dept_slug else []
    attr_cols = [a["attribute_key"] for a in attrs]
    all_cols  = base_cols + ext_cols + attr_cols

    example = {
        "brand": "Brand Name",
        "model_name": "Product Name",
        "model_number": "Model/Part No.",
        "sku": "SKU-001",
        "description": "Short product description",
        "image_url": "https://example.com/image.jpg",
    }
    for col in ext_cols:
        hints = {
            "barcode": "6001234567890", "unit_of_measure": "each|kg|litre|ml|pack",
            "weight_value": "1.5", "weight_unit": "kg|g|lb",
            "shelf_life_days": "730", "regulatory_ref": "A1-1234",
            "requires_rxn": "Yes|No", "dimensions_cm": "120x60x75",
        }
        example[col] = hints.get(col, "")
    for a in attrs:
        example[a["attribute_key"]] = a["unit"] or a["data_type"]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_cols)
    writer.writeheader()
    writer.writerow(example)

    slug = cat["slug"]
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={slug}_template.csv"},
    )


# ── Export products CSV ────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/export.csv")
def catalogue_category_export(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    attrs = _cat_attrs(cur, category_id)

    cur.execute(
        "SELECT * FROM catalogue_products WHERE category_id=%s ORDER BY brand, model_name",
        (category_id,)
    )
    products = cur.fetchall()

    if products:
        pids = [p["id"] for p in products]
        cur.execute(
            """SELECT pa.product_id, ad.attribute_key, pa.value
               FROM catalogue_product_attributes pa
               JOIN catalogue_attribute_definitions ad ON ad.id = pa.attribute_def_id
               WHERE pa.product_id = ANY(%s)""",
            (pids,)
        )
        attr_map: dict = {}
        for row in cur.fetchall():
            attr_map.setdefault(row["product_id"], {})[row["attribute_key"]] = row["value"]
    else:
        attr_map = {}

    cur.close(); conn.close()

    base_cols = ["id", "brand", "model_name", "model_number", "sku", "is_active", "created_at"]
    ext_cols  = ["barcode", "unit_of_measure", "weight_value", "weight_unit",
                 "shelf_life_days", "requires_rxn", "regulatory_ref", "dimensions_cm"]
    attr_cols = [a["attribute_key"] for a in attrs]
    all_cols  = base_cols + ext_cols + attr_cols

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()
    for p in products:
        row = {c: p.get(c, "") for c in base_cols + ext_cols}
        row.update(attr_map.get(p["id"], {}))
        writer.writerow(row)

    slug = cat["slug"]
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={slug}_export.csv"},
    )


# ── Upload CSV/XLSX ────────────────────────────────────────────────────────────

@portal_admin_bp.route("/catalogue/categories/<int:category_id>/upload", methods=["GET", "POST"])
def catalogue_category_upload(category_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cat  = _cat_get(cur, category_id)
    if not cat:
        cur.close(); conn.close()
        flash("Category not found.", "danger")
        return redirect(url_for("portal_admin.catalogue_dashboard"))

    attrs = _cat_attrs(cur, category_id)

    if request.method == "GET":
        cur.close(); conn.close()
        return render_template("portal/admin_category_upload.html", cat=cat, attrs=attrs)

    file = request.files.get("file")
    if not file or not file.filename:
        flash("No file selected.", "danger")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_upload", category_id=category_id))

    on_conflict = request.form.get("on_conflict", "skip")  # "skip" or "update"
    filename    = file.filename.lower()
    inserted = updated = errors = 0
    error_details = []

    required_attrs = [a for a in attrs if a["is_required"]]
    attr_by_key    = {a["attribute_key"]: a for a in attrs}

    try:
        if filename.endswith(".csv"):
            stream    = io.StringIO(file.stream.read().decode("utf-8-sig"))
            rows_iter = list(csv.DictReader(stream))
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb      = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws      = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows_iter = [
                {k: v for k, v in zip(headers, row)}
                for row in ws.iter_rows(min_row=2, values_only=True)
            ]
        else:
            flash("Only .csv or .xlsx files are supported.", "danger")
            cur.close(); conn.close()
            return redirect(url_for("portal_admin.catalogue_category_upload", category_id=category_id))
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_upload", category_id=category_id))

    total_rows = len(rows_iter)

    for row_num, row in enumerate(rows_iter, start=2):
        model_name = str(row.get("model_name") or "").strip()
        if not model_name:
            errors += 1
            error_details.append({"row": row_num, "error": "missing model_name"})
            continue

        # Check required attributes
        missing = [a["attribute_key"] for a in required_attrs
                   if not str(row.get(a["attribute_key"]) or "").strip()]
        if missing:
            errors += 1
            error_details.append({"row": row_num, "error": f"missing required: {', '.join(missing)}"})
            continue

        brand        = str(row.get("brand") or "").strip() or None
        model_number = str(row.get("model_number") or "").strip() or None
        sku_val      = str(row.get("sku") or "").strip() or None
        description  = str(row.get("description") or "").strip() or None
        image_url    = str(row.get("image_url") or "").strip() or None

        # Extended fields from CSV
        barcode         = str(row.get("barcode") or "").strip() or None
        unit_of_measure = str(row.get("unit_of_measure") or "").strip() or None
        wv              = str(row.get("weight_value") or "").strip()
        try:
            weight_value = float(wv) if wv else None
        except ValueError:
            weight_value = None
        weight_unit     = str(row.get("weight_unit") or "").strip() or None
        sl              = str(row.get("shelf_life_days") or "").strip()
        shelf_life_days = int(sl) if sl.isdigit() else None
        rxn_s           = str(row.get("requires_rxn") or "").strip().lower()
        requires_rxn    = True if rxn_s in ("yes","true","1") else (False if rxn_s in ("no","false","0") else None)
        regulatory_ref  = str(row.get("regulatory_ref") or "").strip() or None
        dimensions_cm   = str(row.get("dimensions_cm") or "").strip() or None

        try:
            cur.execute("SAVEPOINT sp_row")
            if on_conflict == "update" and sku_val:
                cur.execute(
                    """INSERT INTO catalogue_products
                       (category_id, brand, model_name, model_number, sku, description, image_url,
                        barcode, unit_of_measure, weight_value, weight_unit,
                        shelf_life_days, requires_rxn, regulatory_ref, dimensions_cm)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (sku) DO UPDATE SET
                         brand=EXCLUDED.brand, model_name=EXCLUDED.model_name,
                         model_number=EXCLUDED.model_number, description=EXCLUDED.description,
                         image_url=EXCLUDED.image_url,
                         barcode=EXCLUDED.barcode, unit_of_measure=EXCLUDED.unit_of_measure,
                         weight_value=EXCLUDED.weight_value, weight_unit=EXCLUDED.weight_unit,
                         shelf_life_days=EXCLUDED.shelf_life_days, requires_rxn=EXCLUDED.requires_rxn,
                         regulatory_ref=EXCLUDED.regulatory_ref, dimensions_cm=EXCLUDED.dimensions_cm,
                         updated_at=NOW()
                       RETURNING id, xmax""",
                    (category_id, brand, model_name, model_number, sku_val, description, image_url,
                     barcode, unit_of_measure, weight_value, weight_unit,
                     shelf_life_days, requires_rxn, regulatory_ref, dimensions_cm)
                )
            else:
                cur.execute(
                    """INSERT INTO catalogue_products
                       (category_id, brand, model_name, model_number, sku, description, image_url,
                        barcode, unit_of_measure, weight_value, weight_unit,
                        shelf_life_days, requires_rxn, regulatory_ref, dimensions_cm)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id, xmax""",
                    (category_id, brand, model_name, model_number, sku_val, description, image_url,
                     barcode, unit_of_measure, weight_value, weight_unit,
                     shelf_life_days, requires_rxn, regulatory_ref, dimensions_cm)
                )

            result = cur.fetchone()
            if not result:
                errors += 1
                cur.execute("ROLLBACK TO SAVEPOINT sp_row")
                continue

            product_id  = result["id"]
            was_updated = result["xmax"] != 0  # xmax!=0 means UPDATE happened

            if was_updated:
                updated += 1
            else:
                inserted += 1

            # Upsert attribute values
            for key, ad in attr_by_key.items():
                val = str(row.get(key) or "").strip() or None
                if val is not None:
                    cur.execute(
                        """INSERT INTO catalogue_product_attributes (product_id, attribute_def_id, value)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (product_id, attribute_def_id) DO UPDATE SET value=EXCLUDED.value""",
                        (product_id, ad["id"], val)
                    )

        except Exception as e:
            errors += 1
            error_details.append({"row": row_num, "error": str(e)})
            cur.execute("ROLLBACK TO SAVEPOINT sp_row")
            continue

        cur.execute("RELEASE SAVEPOINT sp_row")

    try:
        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f"Commit failed: {e}", "danger")
        cur.close(); conn.close()
        return redirect(url_for("portal_admin.catalogue_category_upload", category_id=category_id))

    # Log the upload
    cur.execute(
        """INSERT INTO catalogue_uploads
           (admin_username, category_id, filename, total_rows, successful, failed, status, error_details)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
        (
            _admin_user(), category_id, file.filename,
            total_rows, inserted + updated, errors,
            "partial" if errors > 0 else "completed",
            _json_cat.dumps(error_details),
        )
    )
    conn.commit()
    cur.close(); conn.close()

    insert_audit_log(action="catalogue_upload", admin_username=_admin_user(),
                     details={"category_id": category_id, "filename": file.filename,
                               "inserted": inserted, "updated": updated, "errors": errors})

    msg = f"Import complete — {inserted} added, {updated} updated, {errors} skipped."
    flash(msg, "success" if errors == 0 else "warning")
    return redirect(url_for("portal_admin.catalogue_category_products", category_id=category_id))


# ══════════════════════════════════════════════════════════════════════════════
# PLANS MANAGEMENT — admin view/assign plans to tenants
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/plans")
def admin_plans():
    _require_admin()
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("SELECT * FROM plans ORDER BY sort_order")
    plans = cur.fetchall() or []

    cur.execute("""
        SELECT t.id AS tenant_id, t.name AS business_name,
               t.billing_cycle, t.plan_period_start,
               COALESCE(p.slug, 'free') AS plan_slug,
               COALESCE(p.name, 'Free') AS plan_name,
               (SELECT COUNT(*) FROM usage_events ue
                WHERE ue.tenant_id=t.id AND ue.created_at >= COALESCE(t.plan_period_start, CURRENT_DATE - 30))
                    AS msgs_used,
               COALESCE(p.ai_messages_limit, 100) AS msgs_limit
        FROM tenants t
        LEFT JOIN plans p ON p.id = t.plan_id
        WHERE t.status != 'cancelled'
        ORDER BY t.name
    """)
    tenants = cur.fetchall() or []
    cur.close(); conn.close()

    return render_template("portal/admin_plans.html",
                           plans=plans, tenants=tenants)


@portal_admin_bp.route("/plans/assign/<int:tenant_id>", methods=["POST"])
def admin_plans_assign(tenant_id: int):
    _require_admin()
    plan_id       = int(request.form.get("plan_id") or 1)
    billing_cycle = request.form.get("billing_cycle", "monthly")
    from datetime import date as _d
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("""
        UPDATE tenants
        SET plan_id=%s, billing_cycle=%s, plan_period_start=%s, quota_notified_at=NULL
        WHERE id=%s
    """, (plan_id, billing_cycle, _d.today(), tenant_id))
    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="plan_assign", admin_username=_admin_user(),
                     details={"tenant_id": tenant_id, "plan_id": plan_id, "billing_cycle": billing_cycle})
    flash(f"Plan updated for tenant #{tenant_id}.", "success")
    # Redirect back to wherever the form was submitted from
    referrer = request.referrer or ""
    if "customers" in referrer:
        # Find the customer id for this tenant and go back to their detail page
        try:
            conn2 = get_db_connection()
            cur2  = conn2.cursor()
            cur2.execute("SELECT id FROM customers WHERE tenant_id=%s AND is_active=TRUE ORDER BY id LIMIT 1", (tenant_id,))
            row = cur2.fetchone()
            cur2.close(); conn2.close()
            if row:
                return redirect(url_for("portal_admin.customer_detail", customer_id=row[0]))
        except Exception:
            pass
    return redirect(url_for("portal_admin.admin_plans"))


# ══════════════════════════════════════════════════════════════════════════════
# SCHOOL PLANS MANAGEMENT — admin view/assign plans to schools (school.phixtra.com)
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/school-plans")
def admin_school_plans():
    _require_admin()
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("SELECT * FROM school_plans ORDER BY sort_order")
    plans = cur.fetchall() or []

    cur.execute("""
        SELECT s.id AS school_id, s.school_name,
               s.billing_cycle, s.plan_period_start,
               COALESCE(p.slug, 'free') AS plan_slug,
               COALESCE(p.name, 'Free') AS plan_name,
               (SELECT COUNT(*) FROM school_students st
                WHERE st.school_id=s.id AND st.is_active=TRUE) AS student_count,
               (SELECT COUNT(*) FROM school_chat_history ch
                WHERE ch.school_id=s.id AND ch.role='assistant'
                  AND ch.created_at >= COALESCE(s.plan_period_start, CURRENT_DATE - 30)) AS msgs_used,
               COALESCE(p.ai_messages_limit, 100) AS msgs_limit
        FROM school_profiles s
        LEFT JOIN school_plans p ON p.id = s.plan_id
        ORDER BY s.school_name
    """)
    schools = cur.fetchall() or []
    cur.close(); conn.close()

    return render_template("portal/admin_school_plans.html",
                           plans=plans, schools=schools)


@portal_admin_bp.route("/school-plans/assign/<int:school_id>", methods=["POST"])
def admin_school_plans_assign(school_id: int):
    _require_admin()
    plan_id       = int(request.form.get("plan_id") or 1)
    billing_cycle = request.form.get("billing_cycle", "termly")
    from datetime import date as _d
    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("""
        UPDATE school_profiles
        SET plan_id=%s, billing_cycle=%s, plan_period_start=%s,
            quota_notified_at=NULL, renewal_notified_at=NULL
        WHERE id=%s
    """, (plan_id, billing_cycle, _d.today(), school_id))
    conn.commit()
    cur.close(); conn.close()
    insert_audit_log(action="school_plan_assign", admin_username=_admin_user(),
                     details={"school_id": school_id, "plan_id": plan_id, "billing_cycle": billing_cycle})
    flash(f"Plan updated for school #{school_id}.", "success")
    return redirect(url_for("portal_admin.admin_school_plans"))


# ══════════════════════════════════════════════════════════════════════════════
# WHATSAPP DIAGNOSTICS — admin troubleshooting tool
# ══════════════════════════════════════════════════════════════════════════════

@portal_admin_bp.route("/wa-diagnostics")
def wa_diagnostics():
    _require_admin()
    from datetime import datetime as _dt, timezone as _tz

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT
            wt.id,
            wt.tenant_id,
            wt.phone_number_id,
            wt.display_phone_number,
            wt.verified_name,
            wt.waba_id,
            wt.active,
            wt.token_expires_at,
            wt.signup_method,
            wt.created_at,
            t.name          AS business_name,
            t.status        AS tenant_status,
            MAX(wml.created_at) FILTER (WHERE wml.direction = 'inbound')   AS last_inbound_at,
            MAX(wml.created_at)                                             AS last_any_at,
            COUNT(wml.id)   FILTER (WHERE wml.created_at >= NOW() - INTERVAL '24 hours') AS msgs_24h,
            COUNT(wml.id)   FILTER (WHERE wml.created_at >= NOW() - INTERVAL '7 days')   AS msgs_7d,
            COUNT(DISTINCT wml.customer_phone)
                            FILTER (WHERE wml.created_at >= NOW() - INTERVAL '7 days')   AS customers_7d
        FROM wa_tenants wt
        JOIN tenants t ON t.id = wt.tenant_id
        LEFT JOIN wa_message_log wml ON wml.tenant_id = wt.tenant_id
        GROUP BY wt.id, wt.tenant_id, wt.phone_number_id, wt.display_phone_number,
                 wt.verified_name, wt.waba_id, wt.active, wt.token_expires_at,
                 wt.signup_method, wt.created_at, t.name, t.status
        ORDER BY t.name
    """)
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    now = _dt.now(_tz.utc)

    tenants = []
    for r in rows:
        d = dict(r)

        # Token status
        exp = d.get("token_expires_at")
        if exp is None:
            d["token_status"] = "permanent"
            d["token_label"]  = "Permanent"
        elif exp < now:
            d["token_status"] = "expired"
            d["token_label"]  = f"Expired {exp.strftime('%-d %b %Y')}"
        elif (exp - now).days < 7:
            d["token_status"] = "expiring"
            d["token_label"]  = f"Expires {exp.strftime('%-d %b %Y')}"
        else:
            d["token_status"] = "ok"
            d["token_label"]  = f"OK until {exp.strftime('%-d %b %Y')}"

        # Webhook health (based on last inbound message)
        last = d.get("last_inbound_at")
        if last is None:
            d["webhook_status"] = "none"
            d["webhook_label"]  = "No messages yet"
        else:
            age_h = (now - last).total_seconds() / 3600
            if age_h < 1:
                d["webhook_status"] = "active"
                d["webhook_label"]  = "Active (< 1h ago)"
            elif age_h < 24:
                d["webhook_status"] = "recent"
                d["webhook_label"]  = f"Recent ({int(age_h)}h ago)"
            elif age_h < 168:
                d["webhook_status"] = "quiet"
                d["webhook_label"]  = f"Quiet ({int(age_h/24)}d ago)"
            else:
                d["webhook_status"] = "stale"
                d["webhook_label"]  = f"Stale ({int(age_h/24)}d ago)"

        tenants.append(d)

    return render_template("portal/admin_wa_diagnostics.html", tenants=tenants)


@portal_admin_bp.route("/merchant-signup-qr")
def merchant_signup_qr():
    r = _require_admin()
    if r: return r
    import qrcode
    from flask import send_file
    signup_url = os.getenv("PORTAL_BASE_URL", "https://portal.phixtra.com") + "/register"
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(signup_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    as_dl = request.args.get("download") == "1"
    return send_file(buf, mimetype="image/png",
                     as_attachment=as_dl,
                     download_name="phixtra-merchant-signup-qr.png")


# ── Ambassador Approvals ───────────────────────────────────────────────────

@portal_admin_bp.route("/ambassadors", methods=["GET"])
def ambassadors():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT a.*,
               (SELECT COUNT(*) FROM tenants t WHERE t.ref_code = a.ref_code AND t.status='active') AS active_clients,
               (SELECT COALESCE(SUM(commission_amount),0) FROM ambassador_commissions ac WHERE ac.ambassador_id = a.id) AS total_earned
        FROM ambassadors a
        ORDER BY CASE a.status WHEN 'pending' THEN 0 WHEN 'active' THEN 1 ELSE 2 END, a.created_at DESC
    """)
    ambs = cur.fetchall() or []

    cur.execute("""
        SELECT id, first_name, last_name, ref_code, managed_product FROM ambassadors
        WHERE role='sales_manager' ORDER BY first_name, last_name
    """)
    sales_managers = cur.fetchall() or []

    cur.execute("SELECT ambassador_id, product, status FROM ambassador_products")
    products_by_amb = {}
    for pr in (cur.fetchall() or []):
        products_by_amb.setdefault(pr["ambassador_id"], {})[pr["product"]] = pr["status"]

    cur.close(); conn.close()
    base_url = os.getenv("PORTAL_BASE_URL", "https://portal.phixtra.com")
    return render_template("portal/admin_ambassadors.html", ambassadors=ambs,
                           sales_managers=sales_managers, base_url=base_url,
                           products_by_amb=products_by_amb)


@portal_admin_bp.route("/ambassadors/demo-accounts", methods=["GET"])
def ambassador_demo_accounts():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT a.id AS ambassador_id, a.first_name, a.last_name, a.email AS ambassador_email,
               a.ref_code, a.status AS ambassador_status,
               t.id AS tenant_id, t.name AS tenant_name, t.status AS tenant_status, t.created_at,
               c.id AS customer_id, c.email AS demo_email
        FROM ambassadors a
        JOIN tenants t ON t.id = a.demo_tenant_id
        LEFT JOIN customers c ON c.tenant_id = t.id
        WHERE a.demo_tenant_id IS NOT NULL
        ORDER BY t.created_at DESC
    """)
    demos = cur.fetchall() or []
    cur.close(); conn.close()
    return render_template("portal/admin_ambassador_demos.html", demos=demos)


@portal_admin_bp.route("/ambassadors/<int:amb_id>/qr.png")
def ambassador_qr(amb_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur.close(); conn.close()

    if not amb:
        return "Not found", 404

    base_url = os.getenv("PORTAL_BASE_URL", "https://portal.phixtra.com")
    url = f"{base_url}/register?ref={amb['ref_code']}"

    import qrcode
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    as_dl = request.args.get("download") == "1"
    safe_name = f"{amb['first_name']} {amb['last_name']}".strip().replace(" ", "-").lower()
    return send_file(buf, mimetype="image/png",
                     as_attachment=as_dl,
                     download_name=f"phixtra-qr-{safe_name}.png")


@portal_admin_bp.route("/ambassadors/<int:amb_id>/email-qr", methods=["POST"])
def ambassador_email_qr(amb_id: int):
    r = _require_admin()
    if r: return r

    to_email = (request.form.get("to_email") or "").strip().lower()
    if not to_email:
        return {"error": "No email address provided."}, 400

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur.close(); conn.close()

    if not amb:
        return {"error": "Ambassador not found."}, 404

    base_url = os.getenv("PORTAL_BASE_URL", "https://portal.phixtra.com")
    url      = f"{base_url}/register?ref={amb['ref_code']}"

    import qrcode
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=12, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    png_bytes = buf.getvalue()

    full_name = f"{amb['first_name']} {amb['last_name']}".strip()
    safe_name = full_name.replace(" ", "-").lower()
    BRAND = "#030C18"
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px">
      <h2 style="color:{BRAND}">Your PhiXtra Ambassador QR Code</h2>
      <p>Hi {amb['first_name']},</p>
      <p>Here is your personal QR code. Share it with businesses — when they scan it and sign up,
         they'll be linked to you automatically.</p>
      <p><strong>Your referral link:</strong><br>
         <a href="{url}" style="color:{BRAND}">{url}</a></p>
      <p>The QR code PNG is attached to this email. You can print it, add it to slides,
         or share it digitally.</p>
      <p style="color:#888;font-size:12px">Questions? Contact support@phixtra.com</p>
    </div>"""

    ok = send_email_with_attachment(
        to_email=to_email,
        subject=f"Your PhiXtra Ambassador QR Code — {full_name}",
        html_body=html,
        attachment_bytes=png_bytes,
        attachment_filename=f"phixtra-qr-{safe_name}.png",
        text_body=f"Your PhiXtra referral link: {url}",
    )

    if ok:
        return {"success": True, "message": f"QR code emailed to {to_email}."}, 200
    else:
        return {"error": "Failed to send email. Check SMTP settings."}, 500


@portal_admin_bp.route("/ambassadors/report", methods=["GET"])
def ambassador_report():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT a.id, a.first_name, a.last_name, a.email, a.status,
               (SELECT COUNT(*) FROM tenants t WHERE t.ref_code = a.ref_code AND t.status='active') AS active_clients,
               (SELECT COALESCE(SUM(commission_amount),0) FROM ambassador_commissions ac WHERE ac.ambassador_id = a.id) AS total_earned
        FROM ambassadors a
        ORDER BY CASE a.status WHEN 'pending' THEN 0 WHEN 'active' THEN 1 ELSE 2 END, a.last_name, a.first_name
    """)
    ambs = cur.fetchall() or []
    cur.close(); conn.close()
    counts = {
        "all":       len(ambs),
        "active":    sum(1 for a in ambs if a["status"] == "active"),
        "pending":   sum(1 for a in ambs if a["status"] == "pending"),
        "suspended": sum(1 for a in ambs if a["status"] == "suspended"),
        "rejected":  sum(1 for a in ambs if a["status"] == "rejected"),
    }
    return render_template("portal/admin_ambassador_report.html", ambassadors=ambs, counts=counts)


@portal_admin_bp.route("/ambassadors/report/download", methods=["POST"])
def ambassador_report_download():
    r = _require_admin()
    if r: return r

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date as _date

    amb_ids_raw = request.form.getlist("amb_ids")
    if not amb_ids_raw:
        flash("No ambassadors selected.", "warning")
        return redirect(url_for("portal_admin.ambassador_report"))
    try:
        amb_ids = [int(x) for x in amb_ids_raw]
    except ValueError:
        flash("Invalid selection.", "danger")
        return redirect(url_for("portal_admin.ambassador_report"))

    include_personal  = "col_personal"  in request.form
    include_contact   = "col_contact"   in request.form
    include_location  = "col_location"  in request.form
    include_bank      = "col_bank"      in request.form
    include_id_info   = "col_id"        in request.form
    include_earnings  = "col_earnings"  in request.form
    include_referrals = "col_referrals" in request.form

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    placeholders = ",".join(["%s"] * len(amb_ids))
    cur.execute(f"""
        SELECT a.*,
               (SELECT COUNT(*) FROM tenants t WHERE t.ref_code = a.ref_code AND t.status='active') AS active_clients,
               (SELECT COUNT(*) FROM tenants t WHERE t.ref_code = a.ref_code) AS total_clients,
               (SELECT COALESCE(SUM(commission_amount),0)
                FROM ambassador_commissions ac WHERE ac.ambassador_id = a.id) AS total_earned,
               (SELECT COUNT(*)
                FROM ambassador_commissions ac WHERE ac.ambassador_id = a.id) AS commission_count
        FROM ambassadors a
        WHERE a.id IN ({placeholders})
        ORDER BY a.last_name, a.first_name
    """, amb_ids)
    ambs = cur.fetchall() or []
    cur.close(); conn.close()

    # ── Build column list ───────────────────────────────────────────────────
    headers = ["#", "Full Name", "Status", "Ref Code", "Joined"]
    if include_personal:  headers += ["Date of Birth", "Gender", "Nationality", "Qualification"]
    if include_contact:   headers += ["Email", "Phone", "WhatsApp"]
    if include_location:  headers += ["Address", "Operating Location"]
    if include_bank:      headers += ["Bank Name", "Account Number", "Account Name", "Sort Code", "SWIFT Code"]
    if include_id_info:   headers += ["ID Document Type"]
    if include_earnings:  headers += ["Total Earned (NGN)", "Commission Count"]
    if include_referrals: headers += ["Active Clients", "Total Clients"]

    # ── Workbook ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ambassadors"

    ink_fill    = PatternFill("solid", fgColor="030C18")
    ink_font    = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    thin_side   = Side(style="thin", color="D1D5DB")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    STATUS_FILL = {
        "active":    PatternFill("solid", fgColor="D1FAE5"),
        "pending":   PatternFill("solid", fgColor="FEF9C3"),
        "suspended": PatternFill("solid", fgColor="FEE2E2"),
        "rejected":  PatternFill("solid", fgColor="F3F4F6"),
    }

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(row=1, column=1, value="PhiXtra AI — Ambassador Report")
    tc.font      = Font(bold=True, size=14, color="030C18", name="Calibri")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sc = ws.cell(row=2, column=1,
                 value=f"Generated: {_date.today().strftime('%d %B %Y')}  |  {len(ambs)} ambassador(s) selected")
    sc.font      = Font(size=10, color="6B7280", name="Calibri")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill      = ink_fill
        c.font      = ink_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = cell_border
    ws.row_dimensions[3].height = 30

    # Data rows
    for ridx, amb in enumerate(ambs, 4):
        status = (amb.get("status") or "").lower()
        sfill  = STATUS_FILL.get(status)
        col    = [1]

        def wc(val, fmt=None, center=False):
            c = ws.cell(row=ridx, column=col[0], value=val)
            c.border    = cell_border
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if center else "left")
            if fmt:      c.number_format = fmt
            if sfill and col[0] == 3: c.fill = sfill   # status cell only
            col[0] += 1
            return c

        wc(ridx - 3, center=True)
        wc(f"{amb.get('first_name','')} {amb.get('last_name','')}".strip())
        wc((status or "").title(), center=True)
        wc(amb.get("ref_code") or "", center=True)
        joined = amb.get("created_at")
        wc(joined.strftime("%d %b %Y") if joined else "", center=True)

        if include_personal:
            dob = amb.get("date_of_birth")
            wc(dob.strftime("%d %b %Y") if dob else "", center=True)
            wc(amb.get("gender") or "")
            wc(amb.get("nationality") or "")
            wc(amb.get("highest_qualification") or "")
        if include_contact:
            wc(amb.get("email") or "")
            wc(amb.get("phone") or "")
            wc(amb.get("whatsapp_number") or "")
        if include_location:
            a_cell = wc(amb.get("address") or "")
            a_cell.alignment = Alignment(vertical="center", wrap_text=True)
            wc(amb.get("location") or "")
        if include_bank:
            wc(amb.get("bank_name") or "")
            wc(amb.get("account_number") or "")
            wc(amb.get("account_name") or "")
            wc(amb.get("sort_code") or "")
            wc(amb.get("swift_code") or "")
        if include_id_info:
            wc(amb.get("id_document_type") or "")
        if include_earnings:
            wc(float(amb.get("total_earned") or 0), fmt='#,##0.00', center=True)
            wc(int(amb.get("commission_count") or 0), center=True)
        if include_referrals:
            wc(int(amb.get("active_clients") or 0), center=True)
            wc(int(amb.get("total_clients") or 0), center=True)

        ws.row_dimensions[ridx].height = 20

    # Column widths
    col_widths = [5, 24, 12, 12, 12]
    if include_personal:  col_widths += [14, 10, 14, 22]
    if include_contact:   col_widths += [26, 16, 16]
    if include_location:  col_widths += [32, 20]
    if include_bank:      col_widths += [20, 16, 22, 12, 12]
    if include_id_info:   col_widths += [20]
    if include_earnings:  col_widths += [18, 16]
    if include_referrals: col_widths += [14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"phixtra-ambassadors-{_date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@portal_admin_bp.route("/ambassadors/<int:amb_id>/approve", methods=["POST"])
def ambassador_approve(amb_id: int):
    r = _require_admin()
    if r: return r
    from datetime import date as _date
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    if amb:
        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE ambassadors
               SET status='active', approved_at=NOW(), approved_by=%s, partnership_start=%s
             WHERE id=%s
        """, (_admin_user(), _date.today(), amb_id))
        # This is base KYC/account approval (unlocks login) — it no longer
        # force-creates a Portal row. It only activates whatever product
        # row(s) the ambassador already applied for (via their recruiting
        # sales manager, or a prior admin "assign" action) and are still
        # 'pending'. An ambassador with zero ambassador_products rows (an
        # organic signup with no recruiter) gets none activated here — an
        # admin must separately assign them a product/manager.
        cur2.execute("""
            UPDATE ambassador_products
               SET status='active', partnership_start=%s, approved_at=NOW(), approved_by=%s
             WHERE ambassador_id=%s AND status='pending'
        """, (_date.today(), _admin_user(), amb_id))
        conn.commit()
        cur2.close()
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_approve",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_status": amb['status'], "new_status": "active"},
        )
        # Create demo portal tenant for this ambassador
        try:
            from ambassador_demo import create_ambassador_demo
            create_ambassador_demo(amb['id'], amb['first_name'], amb['ref_code'])
        except Exception as _e:
            print("⚠️ ambassador approval demo tenant creation failed:", _e)
        # Send approval email
        try:
            from ambassador_routes import _send_approved_email
            _send_approved_email(amb['first_name'], amb['email'], amb['ref_code'])
        except Exception as _e:
            print("⚠️ ambassador approval email failed:", _e)
        # Send WhatsApp welcome message
        try:
            from ambassador_routes import _send_whatsapp_welcome
            _send_whatsapp_welcome(amb['first_name'], amb.get('whatsapp_number', ''), amb['ref_code'])
        except Exception as _e:
            print("⚠️ ambassador WhatsApp welcome failed:", _e)
        flash(f"{amb['first_name']} {amb['last_name']} approved as ambassador.", "success")
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/suspend", methods=["POST"])
def ambassador_suspend(amb_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code, status FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassadors SET status='suspended' WHERE id=%s", (amb_id,))
    cur2.execute("""
        UPDATE ambassador_products SET status='suspended' WHERE ambassador_id=%s AND product='portal'
    """, (amb_id,))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_suspend",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_status": amb['status'], "new_status": "suspended"},
        )
    flash("Ambassador suspended.", "warning")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/reactivate", methods=["POST"])
def ambassador_reactivate(amb_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code, status FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassadors SET status='active' WHERE id=%s", (amb_id,))
    cur2.execute("""
        UPDATE ambassador_products SET status='active' WHERE ambassador_id=%s AND product='portal'
    """, (amb_id,))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_reactivate",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_status": amb['status'], "new_status": "active"},
        )
    flash("Ambassador reactivated.", "success")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/reject", methods=["POST"])
def ambassador_reject(amb_id: int):
    r = _require_admin()
    if r: return r
    reason = (request.form.get("reason") or "").strip() or None
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code, status FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassadors SET status='rejected', rejected_at=NOW(), rejected_reason=%s WHERE id=%s",
                (reason, amb_id))
    cur2.execute("""
        UPDATE ambassador_products SET status='rejected', rejected_at=NOW(), rejected_reason=%s
         WHERE ambassador_id=%s AND product='portal'
    """, (reason, amb_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_reject",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_status": amb['status'], "new_status": "rejected",
                     "reason": reason},
        )
    flash("Ambassador application rejected.", "warning")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/terminate", methods=["POST"])
def ambassador_terminate(amb_id: int):
    r = _require_admin()
    if r: return r
    reason = (request.form.get("reason") or "").strip() or None
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code, status FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassadors SET status='terminated', terminated_at=NOW(), terminated_reason=%s WHERE id=%s",
                (reason, amb_id))
    cur2.execute("""
        UPDATE ambassador_products SET status='terminated', terminated_at=NOW(), terminated_reason=%s
         WHERE ambassador_id=%s AND product='portal'
    """, (reason, amb_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_terminate",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_status": amb['status'], "new_status": "terminated",
                     "reason": reason},
        )
    flash("Ambassador terminated.", "danger")
    return redirect(url_for("portal_admin.ambassadors"))


# ── Per-product (School / Estate) approval ──────────────────────────────────
# Portal keeps the routes above (mirrored into ambassador_products by them).
# School/Estate approval only ever touches the ambassador_products row —
# never the top-level ambassadors.status, which gates login and is owned by
# the Portal flow.

@portal_admin_bp.route("/ambassadors/<int:amb_id>/products/<product>/approve", methods=["POST"])
def ambassador_product_approve(amb_id: int, product: str):
    r = _require_admin()
    if r: return r
    if product not in PRODUCT_LABELS:
        flash("Invalid product.", "danger")
        return redirect(url_for("portal_admin.ambassadors"))
    from datetime import date as _date
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    if amb:
        cur2 = conn.cursor()
        cur2.execute("""
            INSERT INTO ambassador_products (ambassador_id, product, status, partnership_start, approved_at, approved_by)
            VALUES (%s, %s, 'active', %s, NOW(), %s)
            ON CONFLICT (ambassador_id, product) DO UPDATE
                SET status='active', partnership_start=EXCLUDED.partnership_start,
                    approved_at=NOW(), approved_by=EXCLUDED.approved_by
        """, (amb_id, product, _date.today(), _admin_user()))
        conn.commit()
        cur2.close()
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_product_approve",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "product": product, "new_status": "active"},
        )
        flash(f"{amb['first_name']} {amb['last_name']} approved to sell {PRODUCT_LABELS[product]}.", "success")
    cur.close(); conn.close()
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/products/<product>/reject", methods=["POST"])
def ambassador_product_reject(amb_id: int, product: str):
    r = _require_admin()
    if r: return r
    if product not in PRODUCT_LABELS:
        flash("Invalid product.", "danger")
        return redirect(url_for("portal_admin.ambassadors"))
    reason = (request.form.get("reason") or "").strip() or None
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("""
        UPDATE ambassador_products SET status='rejected', rejected_at=NOW(), rejected_reason=%s
         WHERE ambassador_id=%s AND product=%s
    """, (reason, amb_id, product))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_product_reject",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "product": product, "new_status": "rejected", "reason": reason},
        )
    flash(f"{PRODUCT_LABELS[product]} application rejected.", "warning")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/products/<product>/suspend", methods=["POST"])
def ambassador_product_suspend(amb_id: int, product: str):
    r = _require_admin()
    if r: return r
    if product not in PRODUCT_LABELS:
        flash("Invalid product.", "danger")
        return redirect(url_for("portal_admin.ambassadors"))
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("""
        UPDATE ambassador_products SET status='suspended' WHERE ambassador_id=%s AND product=%s
    """, (amb_id, product))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_product_suspend",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "product": product, "new_status": "suspended"},
        )
    flash(f"{PRODUCT_LABELS[product]} access suspended.", "warning")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/products/<product>/reactivate", methods=["POST"])
def ambassador_product_reactivate(amb_id: int, product: str):
    r = _require_admin()
    if r: return r
    if product not in PRODUCT_LABELS:
        flash("Invalid product.", "danger")
        return redirect(url_for("portal_admin.ambassadors"))
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("""
        UPDATE ambassador_products SET status='active' WHERE ambassador_id=%s AND product=%s
    """, (amb_id, product))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_product_reactivate",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "product": product, "new_status": "active"},
        )
    flash(f"{PRODUCT_LABELS[product]} access reactivated.", "success")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/edit", methods=["POST"])
def ambassador_edit(amb_id: int):
    r = _require_admin()
    if r: return r
    f = request.form
    conn = get_db_connection()
    cur  = conn.cursor()
    recruited_by_id = (f.get("recruited_by_id") or "").strip()
    cur.execute("""
        UPDATE ambassadors SET
            first_name=%s, last_name=%s, email=%s, phone=%s, whatsapp_number=%s,
            date_of_birth=%s, gender=%s, nationality=%s, address=%s, location=%s,
            highest_qualification=%s, bank_name=%s, account_number=%s,
            account_name=%s, sort_code=%s, swift_code=%s, recruited_by_id=%s
        WHERE id=%s
    """, (
        (f.get("first_name") or "").strip(),
        (f.get("last_name")  or "").strip(),
        (f.get("email")      or "").strip().lower(),
        (f.get("phone")      or "").strip() or None,
        (f.get("whatsapp_number") or "").strip() or None,
        (f.get("date_of_birth")   or "").strip() or None,
        (f.get("gender")     or "").strip() or None,
        (f.get("nationality") or "").strip() or None,
        (f.get("address")    or "").strip() or None,
        (f.get("location")   or "").strip() or None,
        (f.get("highest_qualification") or "").strip() or None,
        (f.get("bank_name")       or "").strip() or None,
        (f.get("account_number")  or "").strip() or None,
        (f.get("account_name")    or "").strip() or None,
        (f.get("sort_code")  or "").strip() or None,
        (f.get("swift_code") or "").strip() or None,
        int(recruited_by_id) if recruited_by_id.isdigit() else None,
        amb_id,
    ))
    # Assigning/changing the sales manager also enrolls this ambassador in
    # that manager's product (matching what registration-via-recruiter-link
    # would have done) — the main path for attaching an organic (no
    # recruiter) signup to a manager after the fact. Only creates the row if
    # missing; never overwrites an existing active/suspended one.
    if recruited_by_id.isdigit():
        cur.execute("SELECT managed_product FROM ambassadors WHERE id=%s AND role='sales_manager'",
                    (int(recruited_by_id),))
        mgr = cur.fetchone()
        if mgr and mgr[0]:
            cur.execute("""
                INSERT INTO ambassador_products (ambassador_id, product, status)
                VALUES (%s, %s, 'pending')
                ON CONFLICT (ambassador_id, product) DO NOTHING
            """, (amb_id, mgr[0]))
    conn.commit()
    cur.close(); conn.close()
    flash("Ambassador details updated.", "success")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/set-role", methods=["POST"])
def ambassador_set_role(amb_id: int):
    r = _require_admin()
    if r: return r
    new_role = "sales_manager" if (request.form.get("role") == "sales_manager") else "ambassador"
    managed_product = (request.form.get("managed_product") or "").strip() or None

    if new_role == "sales_manager":
        if managed_product not in ("portal", "school", "estate"):
            flash("Choose which single product this sales manager will run before saving.", "danger")
            return redirect(url_for("portal_admin.ambassadors"))
    else:
        # Demoting to plain ambassador: a manager is scoped to one product,
        # so the field is meaningless once they're no longer managing a team.
        managed_product = None

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT first_name, last_name, ref_code, role FROM ambassadors WHERE id=%s", (amb_id,))
    amb = cur.fetchone()
    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassadors SET role=%s, managed_product=%s WHERE id=%s",
                 (new_role, managed_product, amb_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    if amb:
        insert_audit_log(
            admin_username=_admin_user(), action="ambassador_set_role",
            details={"ambassador_id": amb_id, "ambassador_name": f"{amb['first_name']} {amb['last_name']}",
                     "ref_code": amb['ref_code'], "old_role": amb['role'], "new_role": new_role,
                     "managed_product": managed_product},
        )
        label = f" ({managed_product})" if managed_product else ""
        flash(f"{amb['first_name']} {amb['last_name']} is now a {new_role.replace('_',' ')}{label}.", "success")
    return redirect(url_for("portal_admin.ambassadors"))


@portal_admin_bp.route("/ambassadors/<int:amb_id>/detail")
def ambassador_detail(amb_id: int):
    r = _require_admin()
    if r: return r

    date_from = (request.args.get("from") or "").strip() or None
    date_to   = (request.args.get("to") or "").strip() or None

    range_clause = ""
    range_params = []
    if date_from:
        range_clause += " AND ac.created_at::date >= %s"
        range_params.append(date_from)
    if date_to:
        range_clause += " AND ac.created_at::date <= %s"
        range_params.append(date_to)

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"""
        SELECT a.*,
               (SELECT COUNT(*) FROM tenants t WHERE t.ref_code=a.ref_code AND t.status='active') AS active_clients,
               (SELECT COALESCE(SUM(ac.commission_amount),0) FROM ambassador_commissions ac
                WHERE ac.ambassador_id=a.id {range_clause}) AS total_earned
        FROM ambassadors a WHERE a.id=%s
    """, [*range_params, amb_id])
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        from flask import jsonify
        return jsonify({"error": "Not found"}), 404

    cur.execute("""
        SELECT admin_username, action, details, created_at
        FROM audit_logs
        WHERE action LIKE 'ambassador_%%'
          AND (details::jsonb->>'ambassador_id')::int = %s
        ORDER BY created_at DESC
        LIMIT 20
    """, (amb_id,))
    history = cur.fetchall() or []

    recruiter_name = None
    if row.get("recruited_by_id"):
        cur.execute("SELECT first_name, last_name FROM ambassadors WHERE id=%s", (row["recruited_by_id"],))
        rec = cur.fetchone()
        if rec:
            recruiter_name = f"{rec['first_name']} {rec['last_name']}"

    cur.execute(f"""
        SELECT id, first_name, last_name, status, ref_code,
               (SELECT COALESCE(SUM(ac.commission_amount),0) FROM ambassador_commissions ac
                WHERE ac.ambassador_id=ambassadors.id {range_clause}) AS total_earned
        FROM ambassadors WHERE recruited_by_id=%s
        ORDER BY first_name, last_name
    """, [*range_params, amb_id])
    recruits = cur.fetchall() or []

    cur.execute(f"""
        SELECT date_trunc('month', ac.created_at) AS month, COALESCE(SUM(ac.commission_amount),0) AS total
        FROM ambassador_commissions ac
        WHERE ac.ambassador_id=%s AND ac.commission_type='override' {range_clause}
        GROUP BY month
        ORDER BY month DESC
        {"" if (date_from or date_to) else "LIMIT 12"}
    """, [amb_id, *range_params])
    monthly_override = cur.fetchall() or []

    cur.execute(f"""
        SELECT t.ref_code AS recruit_ref_code, date_trunc('month', ac.created_at) AS month,
               SUM(ac.commission_amount) AS total
        FROM ambassador_commissions ac
        JOIN tenants t ON t.id = ac.tenant_id
        WHERE ac.ambassador_id=%s AND ac.commission_type='override' {range_clause}
        GROUP BY t.ref_code, month
        ORDER BY month DESC
    """, [amb_id, *range_params])
    per_recruit_rows = cur.fetchall() or []
    cur.close(); conn.close()

    per_recruit_monthly = {}
    for pr in per_recruit_rows:
        per_recruit_monthly.setdefault(pr["recruit_ref_code"], []).append(
            {"month": pr["month"].isoformat(), "total": float(pr["total"] or 0)}
        )

    amb = dict(row)
    for k, v in amb.items():
        if hasattr(v, 'isoformat'):
            amb[k] = v.isoformat()
        elif v is None:
            amb[k] = ""

    status_history = []
    for h in history:
        details = _json.loads(h["details"]) if h["details"] else {}
        status_history.append({
            "admin_username": h["admin_username"],
            "action": h["action"],
            "old_status": details.get("old_status"),
            "new_status": details.get("new_status"),
            "reason": details.get("reason"),
            "created_at": h["created_at"].isoformat() if h["created_at"] else "",
        })
    amb["status_history"] = status_history
    amb["recruiter_name"] = recruiter_name
    amb["recruits"] = [
        {"id": rc["id"], "name": f"{rc['first_name']} {rc['last_name']}", "status": rc["status"],
         "ref_code": rc["ref_code"], "total_earned": float(rc["total_earned"] or 0)}
        for rc in recruits
    ]
    amb["monthly_override"] = [
        {"month": m["month"].isoformat(), "total": float(m["total"] or 0)}
        for m in monthly_override
    ]
    amb["per_recruit_monthly"] = per_recruit_monthly
    amb["date_from"] = date_from or ""
    amb["date_to"] = date_to or ""

    conn2 = get_db_connection()
    cur2  = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur2.execute("SELECT product, status, partnership_start, rejected_reason FROM ambassador_products WHERE ambassador_id=%s", (amb_id,))
    prod_rows = {p["product"]: p for p in (cur2.fetchall() or [])}
    ref_code = row["ref_code"]
    cur2.execute("SELECT COUNT(*) AS n FROM school_profiles WHERE ref_code=%s AND is_active=TRUE", (ref_code,))
    school_active = (cur2.fetchone() or {}).get("n", 0)
    cur2.execute("SELECT COUNT(*) AS n FROM re_tenants WHERE ref_code=%s AND status='active'", (ref_code,))
    estate_active = (cur2.fetchone() or {}).get("n", 0)
    cur2.close(); conn2.close()

    active_counts = {"portal": amb["active_clients"], "school": school_active, "estate": estate_active}
    products = {}
    for p in ("portal", "school", "estate"):
        pr = prod_rows.get(p)
        products[p] = {
            "status": pr["status"] if pr else "not_enrolled",
            "partnership_start": pr["partnership_start"].isoformat() if pr and pr.get("partnership_start") else None,
            "rejected_reason": pr.get("rejected_reason") if pr else None,
            "active_clients": active_counts[p],
        }
    amb["products"] = products

    from flask import jsonify
    return jsonify(amb)


@portal_admin_bp.route("/ambassadors/<int:amb_id>/export-earnings")
def ambassador_export_earnings(amb_id: int):
    r = _require_admin()
    if r: return r

    date_from = (request.args.get("from") or "").strip() or None
    date_to   = (request.args.get("to") or "").strip() or None

    range_clause = ""
    range_params = []
    if date_from:
        range_clause += " AND ac.created_at::date >= %s"
        range_params.append(date_from)
    if date_to:
        range_clause += " AND ac.created_at::date <= %s"
        range_params.append(date_to)

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT ref_code FROM ambassadors WHERE id=%s", (amb_id,))
    amb_row = cur.fetchone()
    if not amb_row:
        cur.close(); conn.close()
        return "Not found", 404

    cur.execute(f"""
        SELECT rec.first_name, rec.last_name, rec.ref_code, rec.status,
               date_trunc('month', ac.created_at) AS month, SUM(ac.commission_amount) AS total
        FROM ambassador_commissions ac
        JOIN tenants t ON t.id = ac.tenant_id
        JOIN ambassadors rec ON rec.ref_code = t.ref_code
        WHERE ac.ambassador_id=%s AND ac.commission_type='override' AND rec.recruited_by_id=%s {range_clause}
        GROUP BY rec.id, rec.first_name, rec.last_name, rec.ref_code, rec.status, month
        ORDER BY month DESC, rec.first_name, rec.last_name
    """, [amb_id, amb_id, *range_params])
    rows = cur.fetchall() or []
    cur.close(); conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Month", "Recruit Name", "Ref Code", "Recruit Status", "Sales Manager Override Earned (NGN)"])
    for row in rows:
        writer.writerow([
            row["month"].strftime("%B %Y"),
            f"{row['first_name']} {row['last_name']}",
            row["ref_code"],
            row["status"],
            f"{float(row['total'] or 0):.2f}",
        ])

    range_label = ""
    if date_from or date_to:
        range_label = f"_{date_from or 'start'}_to_{date_to or 'end'}"
    filename = f"team-earnings-{amb_row['ref_code']}{range_label}.csv"

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@portal_admin_bp.route("/ambassadors/<int:amb_id>/id-doc")
def ambassador_id_doc(amb_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id_document_path FROM ambassadors WHERE id=%s", (amb_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row.get("id_document_path"):
        return "No document found", 404
    import os as _os
    from flask import send_from_directory
    static_dir = _os.path.join(_os.path.dirname(__file__), "static")
    return send_from_directory(static_dir, row["id_document_path"])


@portal_admin_bp.route("/ambassadors/<int:amb_id>/qual-doc")
def ambassador_qual_doc(amb_id: int):
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT qual_document_path FROM ambassadors WHERE id=%s", (amb_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row.get("qual_document_path"):
        return "No qualification document found", 404
    import os as _os
    from flask import send_from_directory
    static_dir = _os.path.join(_os.path.dirname(__file__), "static")
    return send_from_directory(static_dir, row["qual_document_path"])



# ── Ambassador/Sales Manager CRM Pipeline (admin oversight) ─────────────────

@portal_admin_bp.route("/admin/leads", methods=["GET"])
def admin_leads():
    r = _require_admin()
    if r: return r
    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT al.*,
               a.first_name || ' ' || a.last_name AS ambassador_name,
               a.email AS ambassador_email
        FROM ambassador_leads al
        JOIN ambassadors a ON a.id = al.ambassador_id
        WHERE al.dropped_at IS NULL
        ORDER BY CASE al.stage
            WHEN 'lead' THEN 0 WHEN 'contacted' THEN 1 WHEN 'demo_done' THEN 2
            WHEN 'requirements_confirmed' THEN 3 WHEN 'onboarding' THEN 4
            WHEN 'active_client' THEN 5 WHEN 'support' THEN 6 ELSE 7 END,
            al.created_at DESC
    """)
    leads = [dict(r) for r in cur.fetchall()]

    cur.execute("""
        SELECT al.*, a.first_name || ' ' || a.last_name AS ambassador_name
        FROM ambassador_leads al
        JOIN ambassadors a ON a.id = al.ambassador_id
        WHERE al.dropped_at IS NOT NULL
        ORDER BY al.dropped_at DESC
    """)
    dropped_leads = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()

    stage_counts = {s: 0 for s in STAGE_ORDER}
    for l in leads:
        stage_counts[l["stage"]] = stage_counts.get(l["stage"], 0) + 1

    conn2 = get_db_connection()
    cur2  = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur2.execute("""
        SELECT id, first_name, last_name, ref_code FROM ambassadors
        WHERE status='active' ORDER BY first_name, last_name
    """)
    ambassadors = cur2.fetchall() or []
    cur2.close(); conn2.close()

    return render_template("portal/admin_leads.html", leads=leads, dropped_leads=dropped_leads,
        stage_counts=stage_counts, stage_order=STAGE_ORDER, stage_labels=STAGE_LABELS,
        stage_descriptions=STAGE_DESCRIPTIONS, next_stage=next_stage, ambassadors=ambassadors,
        product_labels=PRODUCT_LABELS, product_icons=PRODUCT_ICONS)


@portal_admin_bp.route("/admin/leads/create", methods=["POST"])
def admin_lead_create():
    r = _require_admin()
    if r: return r

    ambassador_id_raw = (request.form.get("ambassador_id") or "").strip()
    business_name     = (request.form.get("business_name") or "").strip()
    industry          = (request.form.get("industry")      or "").strip() or None
    contact_name      = (request.form.get("contact_name")  or "").strip() or None
    phone             = (request.form.get("phone")         or "").strip() or None
    email             = (request.form.get("email")          or "").strip() or None
    notes             = (request.form.get("notes")          or "").strip() or None
    product           = (request.form.get("product")        or "portal").strip().lower()
    if product not in PRODUCT_LABELS:
        product = "portal"

    if not ambassador_id_raw.isdigit() or not business_name:
        flash("Ambassador and business name are required.", "danger")
        return redirect(url_for("portal_admin.admin_leads"))
    ambassador_id = int(ambassador_id_raw)

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM ambassadors WHERE id=%s", (ambassador_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        flash("Ambassador not found.", "danger")
        return redirect(url_for("portal_admin.admin_leads"))

    cur.execute("""
        INSERT INTO ambassador_leads
          (ambassador_id, business_name, industry, contact_name, phone, email, notes, stage, product)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'lead', %s)
        RETURNING id
    """, (ambassador_id, business_name, industry, contact_name, phone, email, notes, product))
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close(); conn.close()

    record_stage_change(new_id, None, "lead", f"{_admin_user()} (Admin)")
    flash(f"{business_name} added to the pipeline.", "success")
    return redirect(url_for("portal_admin.admin_leads"))


@portal_admin_bp.route("/admin/leads/<int:lead_id>/advance", methods=["POST"])
def lead_advance(lead_id: int):
    r = _require_admin()
    if r: return r

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM ambassador_leads WHERE id=%s", (lead_id,))
    lead = cur.fetchone()
    if not lead:
        cur.close(); conn.close()
        flash("Lead not found.", "danger")
        return redirect(url_for("portal_admin.admin_leads"))

    target = next_stage(lead["stage"])
    if not target:
        cur.close(); conn.close()
        flash("This lead is already at the final stage.", "warning")
        return redirect(url_for("portal_admin.admin_leads"))

    f = request.form
    updates = {"stage": target}

    if target == "contacted":
        contact_channel  = (f.get("contact_channel") or "").strip()
        contact_date     = (f.get("contact_date") or "").strip()
        contact_response = (f.get("contact_response") or "").strip()
        if not contact_channel or not contact_date:
            cur.close(); conn.close()
            flash("Contact channel and date are required.", "danger")
            return redirect(url_for("portal_admin.admin_leads"))
        updates.update(contact_channel=contact_channel, contact_date=contact_date,
                       contact_response=contact_response or None)

    elif target == "demo_done":
        demo_date     = (f.get("demo_date") or "").strip()
        demo_reaction = (f.get("demo_reaction") or "").strip()
        if not demo_date:
            cur.close(); conn.close()
            flash("Demo date is required.", "danger")
            return redirect(url_for("portal_admin.admin_leads"))
        updates.update(demo_date=demo_date, demo_reaction=demo_reaction or None)

    elif target == "requirements_confirmed":
        req_phone    = f.get("req_phone") == "1"
        req_meta     = f.get("req_meta_account") == "1"
        req_whatsapp = f.get("req_whatsapp_connected") == "1"
        req_products = f.get("req_product_list") == "1"
        if not (req_phone and req_meta and req_whatsapp and req_products):
            cur.close(); conn.close()
            flash("All 4 requirements must be confirmed before advancing.", "danger")
            return redirect(url_for("portal_admin.admin_leads"))
        updates.update(req_phone=True, req_meta_account=True,
                       req_whatsapp_connected=True, req_product_list=True)

    elif target == "onboarding":
        onboarding_date  = (f.get("onboarding_date") or "").strip()
        onboarding_notes = (f.get("onboarding_notes") or "").strip()
        if not onboarding_date:
            cur.close(); conn.close()
            flash("Onboarding date is required.", "danger")
            return redirect(url_for("portal_admin.admin_leads"))
        updates.update(onboarding_date=onboarding_date, onboarding_notes=onboarding_notes or None)

    elif target == "active_client":
        lead_product = lead.get("product") or "portal"
        link_col     = LEAD_LINK_COL[lead_product]
        ref_table    = LEAD_REF_TABLE[lead_product]
        tenant_id_raw = (f.get("tenant_id") or "").strip()
        tenant_id = int(tenant_id_raw) if tenant_id_raw.isdigit() else None
        if tenant_id:
            cur.execute("SELECT a.ref_code FROM ambassadors a WHERE a.id=%s", (lead["ambassador_id"],))
            amb_row = cur.fetchone()
            cur.execute(f"SELECT id FROM {ref_table} WHERE id=%s AND ref_code=%s",
                       (tenant_id, amb_row["ref_code"] if amb_row else None))
            if not cur.fetchone():
                cur.close(); conn.close()
                flash("Selected client doesn't match this ambassador's referral code.", "danger")
                return redirect(url_for("portal_admin.admin_leads"))
        updates[link_col] = tenant_id

    set_clause = ", ".join(f"{k}=%s" for k in updates)
    cur2 = conn.cursor()
    cur2.execute(f"UPDATE ambassador_leads SET {set_clause} WHERE id=%s",
                list(updates.values()) + [lead_id])
    conn.commit()
    cur2.close(); cur.close(); conn.close()

    record_stage_change(lead_id, lead["stage"], target, f"{_admin_user()} (Admin)")
    flash(f"{lead['business_name']} moved to {STAGE_LABELS[target]}.", "success")
    return redirect(url_for("portal_admin.admin_leads"))


@portal_admin_bp.route("/admin/leads/<int:lead_id>/drop", methods=["POST"])
def lead_drop_admin(lead_id: int):
    r = _require_admin()
    if r: return r
    reason = (request.form.get("reason") or "").strip() or None

    conn = get_db_connection()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM ambassador_leads WHERE id=%s", (lead_id,))
    lead = cur.fetchone()
    if not lead:
        cur.close(); conn.close()
        flash("Lead not found.", "danger")
        return redirect(url_for("portal_admin.admin_leads"))

    cur2 = conn.cursor()
    cur2.execute("UPDATE ambassador_leads SET dropped_at=NOW(), dropped_reason=%s WHERE id=%s",
                (reason, lead_id))
    conn.commit()
    cur2.close(); cur.close(); conn.close()
    record_stage_change(lead_id, lead["stage"], "dropped", f"{_admin_user()} (Admin)", reason)
    flash(f"{lead['business_name']} marked as dropped.", "warning")
    return redirect(url_for("portal_admin.admin_leads"))


@portal_admin_bp.route("/video-tutorials")
def video_tutorials():
    r = _require_admin()
    if r: return r
    return render_template("portal/admin_video_tutorials.html", videos=TUTORIAL_VIDEOS)


@portal_admin_bp.route("/admin/leads/<int:lead_id>/history")
def lead_history_admin(lead_id: int):
    r = _require_admin()
    if r: return r
    history = get_stage_history(lead_id)
    from flask import jsonify
    return jsonify([
        {"from_stage": h["from_stage"], "to_stage": h["to_stage"], "changed_by": h["changed_by"],
         "notes": h["notes"], "created_at": h["created_at"].isoformat() if h["created_at"] else ""}
        for h in history
    ])
